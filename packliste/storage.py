import json
import pickle
import time
from pathlib import Path
from typing import Sequence

import openpyxl
import pandas as pd


def _warn(logger, msg: str, *args) -> None:
    if logger is not None:
        logger.warning(msg, *args)


def save_last_active_path(meta_file: Path, path: str, *, logger=None) -> None:
    """Persist the last-used Excel path so refreshes can restore it."""
    try:
        with open(meta_file, "w", encoding="utf-8") as fh:
            json.dump({"path": path, "ts": time.time()}, fh)
    except Exception as exc:
        _warn(logger, "Cannot write meta %s: %s", meta_file, exc)


def load_last_active_path(meta_file: Path, *, logger=None) -> str:
    """Return last-active path, or an empty string when unavailable."""
    try:
        if meta_file.exists():
            with open(meta_file, "r", encoding="utf-8") as fh:
                data = json.load(fh)
            path = data.get("path", "")
            if path and Path(path).exists():
                return path
    except Exception as exc:
        _warn(logger, "Cannot read meta %s: %s", meta_file, exc)
    return ""


def _sidecar_path(excel_path: str) -> Path:
    path = Path(excel_path)
    return path.parent / f"{path.stem}_autosave.pkl"


def _write_sidecar(df: pd.DataFrame, excel_path: str) -> None:
    """Atomic synchronous pickle write. Raises on failure."""
    sidecar = _sidecar_path(excel_path)
    tmp = sidecar.with_suffix(".pkl.tmp")
    payload = {"df": df.copy(), "path": excel_path, "ts": time.time()}
    with open(tmp, "wb") as fh:
        pickle.dump(payload, fh, protocol=pickle.HIGHEST_PROTOCOL)
    tmp.replace(sidecar)


def _read_sidecar(excel_path: str, *, logger=None):
    sidecar = _sidecar_path(excel_path)
    if not sidecar.exists():
        return None
    try:
        with open(sidecar, "rb") as fh:
            data = pickle.load(fh)
        if data.get("path") != excel_path:
            return None
        return data["df"]
    except Exception as exc:
        _warn(logger, "Cannot read sidecar %s: %s", sidecar, exc)
        return None


def _sidecar_newer(excel_path: str) -> bool:
    sidecar = _sidecar_path(excel_path)
    excel = Path(excel_path)
    if not sidecar.exists():
        return False
    if not excel.exists():
        return True
    return sidecar.stat().st_mtime > excel.stat().st_mtime


def _detect_sheet_layout(ws, *, required_cols: Sequence[str]) -> tuple[int, int, dict[str, int]]:
    """Return header row index for pandas, first data row in Excel, and header map."""
    for excel_row in range(1, min(ws.max_row, 12) + 1):
        col_map = {
            str(cell.value).strip(): cell.column
            for cell in ws[excel_row]
            if cell.value and str(cell.value).strip()
        }
        if all(col in col_map for col in required_cols):
            data_start_row = excel_row + 1
            for candidate_row in range(excel_row + 1, ws.max_row + 1):
                has_data = any(
                    ws.cell(candidate_row, xl_col).value not in (None, "")
                    for xl_col in col_map.values()
                )
                if has_data:
                    data_start_row = candidate_row
                    break
            return excel_row - 1, data_start_row, col_map

    raise ValueError(
        "Headerzeile im Sheet 'Packliste' nicht gefunden. "
        "Bitte pruefen, ob die Pflichtspalten vorhanden sind."
    )


def load_df(
    path: str,
    *,
    sheet_name: str,
    required_cols: Sequence[str],
    checkbox_cols: Sequence[str],
    checked_value: str,
) -> pd.DataFrame:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=False)
    try:
        worksheet = workbook[sheet_name]
        header_row_idx, _, _ = _detect_sheet_layout(worksheet, required_cols=required_cols)
    finally:
        workbook.close()

    df = pd.read_excel(path, sheet_name=sheet_name, header=header_row_idx, dtype=str)
    df = df.where(df.notna() & (df != "nan"), "")
    missing = [col for col in required_cols if col not in df.columns]
    if missing:
        raise ValueError(
            f"Fehlende Spalten: {missing}\n\n"
            f"Gefundene Spalten: {list(df.columns)}\n\n"
            f"Bitte pruefen, ob die richtige Datei und das Sheet '{sheet_name}' geladen wird."
        )
    for col in checkbox_cols:
        if col in df.columns:
            df[col] = df[col] == checked_value
    if "Verladen" not in df.columns:
        df["Verladen"] = False
    return df


def save_df(
    df: pd.DataFrame,
    path: str,
    *,
    sheet_name: str,
    required_cols: Sequence[str],
    checkbox_cols: Sequence[str],
    checked_value: str,
    unchecked_value: str,
) -> None:
    """
    Near-atomic Excel write:
      1. Write to <stem>.tmp.xlsx
      2. Rename to <path>  (replaces original only after full successful write)
    If anything fails the original file is untouched; tmp is cleaned up.
    """
    target = Path(path)
    tmp = target.parent / (target.stem + ".tmp.xlsx")
    try:
        workbook = openpyxl.load_workbook(path)
        worksheet = workbook[sheet_name]
        header_row_idx, data_start_row, col_map = _detect_sheet_layout(
            worksheet,
            required_cols=required_cols,
        )
        header_row = header_row_idx + 1
        if "Verladen" not in col_map:
            next_col = max(col_map.values()) + 1
            worksheet.cell(row=header_row, column=next_col, value="Verladen")
            col_map["Verladen"] = next_col
        for row_offset, (_, row) in enumerate(df.iterrows()):
            excel_row = data_start_row + row_offset
            for col_name, excel_col in col_map.items():
                if col_name not in df.columns:
                    continue
                value = row[col_name]
                if col_name in checkbox_cols:
                    value = checked_value if bool(value) else unchecked_value
                else:
                    value = value if value != "" else None
                worksheet.cell(row=excel_row, column=excel_col, value=value)
        first_clear_row = data_start_row + len(df)
        if first_clear_row <= worksheet.max_row:
            for excel_row in range(first_clear_row, worksheet.max_row + 1):
                for excel_col in set(col_map.values()):
                    worksheet.cell(row=excel_row, column=excel_col, value=None)
        workbook.save(str(tmp))
        tmp.replace(target)
    except Exception:
        try:
            tmp.unlink(missing_ok=True)
        except Exception:
            pass
        raise


def restore_state_for_path(
    path: str,
    *,
    sheet_name: str,
    required_cols: Sequence[str],
    checkbox_cols: Sequence[str],
    checked_value: str,
    logger=None,
) -> tuple[pd.DataFrame, str]:
    """
    Load best available state for *path*.
    Returns (df, source) where source is "sidecar" or "excel".
    Raises ValueError / IOError on failure.
    """
    if _sidecar_newer(path):
        sidecar_df = _read_sidecar(path, logger=logger)
        if sidecar_df is not None:
            return sidecar_df, "sidecar"
    df = load_df(
        path,
        sheet_name=sheet_name,
        required_cols=required_cols,
        checkbox_cols=checkbox_cols,
        checked_value=checked_value,
    )
    return df, "excel"
