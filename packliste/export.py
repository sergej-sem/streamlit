from io import BytesIO
from pathlib import Path
from typing import Sequence

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def to_excel_bytes(
    df: pd.DataFrame,
    *,
    sheet_name: str = "Liste",
    checkbox_cols: Sequence[str],
) -> bytes:
    data = BytesIO()
    export_df = df.reset_index(drop=True).copy()
    checkbox_cols = set(checkbox_cols)

    for col in checkbox_cols:
        if col in export_df.columns:
            export_df[col] = export_df[col].map(lambda v: "Ja" if bool(v) else "Nein")

    with pd.ExcelWriter(data, engine="openpyxl") as writer:
        export_df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.book[sheet_name]

        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style="thin", color="D9E2F3"),
            right=Side(style="thin", color="D9E2F3"),
            top=Side(style="thin", color="D9E2F3"),
            bottom=Side(style="thin", color="D9E2F3"),
        )
        stripe_fill = PatternFill("solid", fgColor="F7F9FC")

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        ws.sheet_view.showGridLines = False
        ws.row_dimensions[1].height = 24

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border

        for row_idx in range(2, ws.max_row + 1):
            is_striped = row_idx % 2 == 0
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = border
                cell.alignment = cell_alignment
                if is_striped:
                    cell.fill = stripe_fill

        for col_idx, column_name in enumerate(export_df.columns, start=1):
            col_letter = get_column_letter(col_idx)
            values = [column_name, *export_df.iloc[:, col_idx - 1].tolist()]
            max_length = max(len(str(v)) if v is not None else 0 for v in values)
            width = min(max(max_length + 2, 12), 42)
            if column_name in {"Beschreibung", "Notizen"}:
                width = min(max(max_length + 2, 24), 56)
            if column_name in checkbox_cols or column_name == "Fortschritt":
                width = min(max(max_length + 2, 12), 16)
                for row_idx in range(2, ws.max_row + 1):
                    ws.cell(row=row_idx, column=col_idx).alignment = center_alignment
            ws.column_dimensions[col_letter].width = width

        if "Fortschritt" in export_df.columns:
            progress_idx = export_df.columns.get_loc("Fortschritt") + 1
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=progress_idx).number_format = '0" %"'

    data.seek(0)
    return data.getvalue()


def download_filename(pl_path: str | None, slug: str) -> str:
    path = Path(pl_path or "packliste.xlsx")
    return f"{path.stem}_{slug}.xlsx"
