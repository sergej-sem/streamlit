from __future__ import annotations

import re
import subprocess
import tempfile
import textwrap
import unicodedata
from collections.abc import Callable
from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from sponsor_deadline_mails.parser import build_sponsor_row, normalize_text
from shared.mail_message import MailAttachment
from shared.mail_rich_text import render_final_mail_html
from serienmailing.imap_sender import SerienMail
from serienmailing.mail_builder import build_subject


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SHEET_NAME = "Deals"
DEFAULT_WORKBOOK_PATH = PROJECT_ROOT / "00_Master_Sponsoren_Infos.xlsx"
DEFAULT_KONTAKTLISTE_PATH = PROJECT_ROOT / "Kontaktliste.xlsx"
DEFAULT_GESPRAECHSPLAENE_DIR = PROJECT_ROOT / "Gespr\u00e4chspl\u00e4ne"
DEFAULT_VORTRAGSLISTEN_DIR = PROJECT_ROOT / "Vortragslisten"
DEFAULT_KONTAKTLISTE_PASSWORD = "MSEukf6dd"

SUPPORTED_PACKAGES = {"premium": "Premium", "gold": "Gold", "platin": "Platin"}
PREMIUM_KEY = "premium"
GOLD_KEY = "gold"
PLATIN_KEY = "platin"

_TRANSLITERATION_MAP = str.maketrans(
    {
        "\u00e4": "ae",
        "\u00f6": "oe",
        "\u00fc": "ue",
        "\u00df": "ss",
        "\u00c4": "Ae",
        "\u00d6": "Oe",
        "\u00dc": "Ue",
    }
)
_GESPRAECHSPLAN_SUFFIXES = ("gespraechsplan", "gesprachsplan")
_VORTRAGSLISTE_SUFFIXES = ("vortragsliste",)


@dataclass(frozen=True)
class EventEndSourceStatus:
    workbook_path: Path
    workbook_exists: bool
    kontaktliste_path: Path
    kontaktliste_exists: bool
    gespraechsplaene_dir: Path
    gespraechsplaene_exists: bool
    gespraechsplaene_count: int
    vortragslisten_dir: Path
    vortragslisten_exists: bool
    vortragslisten_count: int


@dataclass(frozen=True)
class EventEndSponsorPlan:
    row_number: int
    sponsor_name: str
    package: str
    language: str
    to_email: str
    cc_email: str
    contact_first_name: str
    contact_last_name: str
    attachments: tuple[MailAttachment, ...]
    status: str
    details: str

    @property
    def is_ready(self) -> bool:
        return self.status == "ready"

    @property
    def attachment_names(self) -> tuple[str, ...]:
        return tuple(attachment.filename for attachment in self.attachments)


@dataclass(frozen=True)
class EventEndAssemblyResult:
    sponsors: tuple[EventEndSponsorPlan, ...]
    ready_count: int
    blocked_count: int
    skipped_count: int
    kontaktliste_protected: bool
    kontaktliste_protection_details: str


@dataclass(frozen=True)
class _LoadedAttachmentFile:
    name: str
    content: bytes
    ext: str
    stem_key: str


def normalize_attachment_key(value: str) -> str:
    text = normalize_text(value).translate(_TRANSLITERATION_MAP)
    text = (
        unicodedata.normalize("NFKD", text)
        .encode("ascii", "ignore")
        .decode("ascii")
        .casefold()
    )
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return re.sub(r"_+", "_", text).strip("_")


def inspect_eventende_sources(
    *,
    workbook_path: Path = DEFAULT_WORKBOOK_PATH,
    kontaktliste_path: Path = DEFAULT_KONTAKTLISTE_PATH,
    gespraechsplaene_dir: Path = DEFAULT_GESPRAECHSPLAENE_DIR,
    vortragslisten_dir: Path = DEFAULT_VORTRAGSLISTEN_DIR,
) -> EventEndSourceStatus:
    return EventEndSourceStatus(
        workbook_path=workbook_path,
        workbook_exists=workbook_path.is_file(),
        kontaktliste_path=kontaktliste_path,
        kontaktliste_exists=kontaktliste_path.is_file(),
        gespraechsplaene_dir=gespraechsplaene_dir,
        gespraechsplaene_exists=gespraechsplaene_dir.is_dir(),
        gespraechsplaene_count=_count_files(gespraechsplaene_dir),
        vortragslisten_dir=vortragslisten_dir,
        vortragslisten_exists=vortragslisten_dir.is_dir(),
        vortragslisten_count=_count_files(vortragslisten_dir),
    )


def _count_files(path: Path) -> int:
    if not path.is_dir():
        return 0
    return sum(1 for file_path in path.iterdir() if file_path.is_file())


def _powershell_quote(value: str) -> str:
    return "'" + str(value).replace("'", "''") + "'"


def _protect_excel_file_with_excel_com(source_path: Path, target_path: Path, password: str) -> None:
    script = textwrap.dedent(
        f"""
        $ErrorActionPreference = 'Stop'
        $src = {_powershell_quote(str(source_path))}
        $dst = {_powershell_quote(str(target_path))}
        $password = {_powershell_quote(password)}
        $excel = $null
        $workbook = $null
        try {{
            $excel = New-Object -ComObject Excel.Application
            $excel.Visible = $false
            $excel.DisplayAlerts = $false
            $workbook = $excel.Workbooks.Open($src)
            $workbook.SaveAs($dst, 51, $password)
        }} finally {{
            if ($workbook -ne $null) {{
                try {{ $workbook.Close($false) | Out-Null }} catch {{}}
                [System.Runtime.Interopservices.Marshal]::ReleaseComObject($workbook) | Out-Null
            }}
            if ($excel -ne $null) {{
                try {{ $excel.Quit() | Out-Null }} catch {{}}
                [System.Runtime.Interopservices.Marshal]::ReleaseComObject($excel) | Out-Null
            }}
        }}
        """
    ).strip()
    result = subprocess.run(
        ["powershell", "-NoProfile", "-Command", script],
        capture_output=True,
        text=True,
        check=False,
    )
    if result.returncode != 0:
        detail = (result.stderr or result.stdout or "").strip() or "Unbekannter Excel-Fehler."
        raise RuntimeError(detail)


def build_password_protected_excel_attachment(
    source_path: Path,
    password: str = DEFAULT_KONTAKTLISTE_PASSWORD,
    *,
    protector: Callable[[Path, Path, str], None] | None = None,
) -> MailAttachment:
    if not source_path.is_file():
        raise FileNotFoundError(f"Kontaktliste nicht gefunden: {source_path}")

    protect = protector or _protect_excel_file_with_excel_com
    with tempfile.TemporaryDirectory() as tmpdir:
        protected_path = Path(tmpdir) / source_path.name
        protect(source_path, protected_path, password)
        if not protected_path.is_file():
            raise RuntimeError("Die verschlüsselte Kontaktliste wurde nicht erzeugt.")
        return MailAttachment(filename=source_path.name, content=protected_path.read_bytes())


def _load_attachment_dir(path: Path) -> tuple[_LoadedAttachmentFile, ...]:
    if not path.is_dir():
        return ()
    files = []
    for file_path in sorted(path.iterdir(), key=lambda item: item.name.casefold()):
        if not file_path.is_file():
            continue
        files.append(
            _LoadedAttachmentFile(
                name=file_path.name,
                content=file_path.read_bytes(),
                ext=file_path.suffix.lower(),
                stem_key=normalize_attachment_key(file_path.stem),
            )
        )
    return tuple(files)


def _package_key(value: str) -> str:
    return normalize_text(value).casefold()


def _strip_suffix(stem_key: str, suffixes: tuple[str, ...]) -> str:
    for suffix in suffixes:
        if stem_key == suffix:
            return ""
        marker = f"_{suffix}"
        if stem_key.endswith(marker):
            return stem_key[: -len(marker)]
    return stem_key


def _tokens(key: str) -> list[str]:
    return [token for token in key.split("_") if token]


def _tokens_match_prefix(shorter: list[str], longer: list[str]) -> bool:
    if not shorter or len(shorter) > len(longer):
        return False
    return all(
        longer[index].startswith(shorter[index]) or shorter[index].startswith(longer[index])
        for index in range(len(shorter))
    )


def _keys_match(sponsor_key: str, file_key: str) -> bool:
    if not sponsor_key or not file_key:
        return False
    if sponsor_key == file_key:
        return True
    if sponsor_key.startswith(file_key + "_") or file_key.startswith(sponsor_key + "_"):
        return True

    sponsor_tokens = _tokens(sponsor_key)
    file_tokens = _tokens(file_key)
    if len(sponsor_tokens) <= len(file_tokens):
        return _tokens_match_prefix(sponsor_tokens, file_tokens)
    return _tokens_match_prefix(file_tokens, sponsor_tokens)


def _matching_attachments(
    sponsor_name: str,
    files: tuple[_LoadedAttachmentFile, ...],
    *,
    suffixes: tuple[str, ...],
    allowed_exts: tuple[str, ...],
) -> tuple[MailAttachment, ...]:
    sponsor_key = normalize_attachment_key(sponsor_name)
    matches: list[MailAttachment] = []
    for file in files:
        if file.ext not in allowed_exts:
            continue
        file_key = _strip_suffix(file.stem_key, suffixes)
        if _keys_match(sponsor_key, file_key):
            matches.append(MailAttachment(filename=file.name, content=file.content))
    matches.sort(key=lambda attachment: attachment.filename.casefold())
    return tuple(matches)


def assemble_eventende_sponsors(
    *,
    workbook_path: Path = DEFAULT_WORKBOOK_PATH,
    kontaktliste_path: Path = DEFAULT_KONTAKTLISTE_PATH,
    gespraechsplaene_dir: Path = DEFAULT_GESPRAECHSPLAENE_DIR,
    vortragslisten_dir: Path = DEFAULT_VORTRAGSLISTEN_DIR,
    sheet_name: str = DEFAULT_SHEET_NAME,
    kontaktliste_password: str = DEFAULT_KONTAKTLISTE_PASSWORD,
    kontaktliste_attachment_builder: Callable[[Path, str], MailAttachment] | None = None,
) -> EventEndAssemblyResult:
    if not workbook_path.is_file():
        raise FileNotFoundError(f"Sponsoren-Datei nicht gefunden: {workbook_path}")

    kontaktliste: MailAttachment | None = None
    kontaktliste_protected = False
    kontaktliste_protection_details = ""
    builder = kontaktliste_attachment_builder or build_password_protected_excel_attachment
    if not kontaktliste_path.is_file():
        kontaktliste_protection_details = "Kontaktliste fehlt."
    else:
        try:
            kontaktliste = builder(kontaktliste_path, kontaktliste_password)
            kontaktliste_protected = True
            kontaktliste_protection_details = "Kontaktliste erfolgreich verschlüsselt."
        except Exception as exc:
            kontaktliste_protection_details = f"Kontaktliste-Verschlüsselung fehlgeschlagen: {exc}"

    gespraechsplaene = _load_attachment_dir(gespraechsplaene_dir)
    vortragslisten = _load_attachment_dir(vortragslisten_dir)

    workbook = load_workbook(workbook_path, data_only=False)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"Blatt '{sheet_name}' nicht gefunden. Verf\u00fcgbar: {', '.join(workbook.sheetnames)}"
            )

        ws = workbook[sheet_name]
        sponsors: list[EventEndSponsorPlan] = []
        candidate_rows = 0

        for row_number in range(2, ws.max_row + 1):
            sponsor_name = normalize_text(ws[f"B{row_number}"].value)
            if not sponsor_name:
                continue
            candidate_rows += 1
            sponsor = build_sponsor_row(ws, row_number)
            if sponsor is None:
                continue

            package_key = _package_key(sponsor.package)
            if package_key not in SUPPORTED_PACKAGES:
                continue

            issues: list[str] = []
            attachments: list[MailAttachment] = []

            if not kontaktliste_protected or kontaktliste is None:
                issues.append(kontaktliste_protection_details or "Kontaktliste fehlt.")
            else:
                attachments.append(kontaktliste)

            matched_gespraechsplaene = _matching_attachments(
                sponsor.sponsor_name,
                gespraechsplaene,
                suffixes=_GESPRAECHSPLAN_SUFFIXES,
                allowed_exts=(".pdf",),
            )
            if not matched_gespraechsplaene:
                issues.append("Gespr\u00e4chsplan fehlt.")
            else:
                attachments.extend(matched_gespraechsplaene)

            if package_key in {GOLD_KEY, PLATIN_KEY}:
                matched_vortragslisten = _matching_attachments(
                    sponsor.sponsor_name,
                    vortragslisten,
                    suffixes=_VORTRAGSLISTE_SUFFIXES,
                    allowed_exts=(".pdf",),
                )
                if not matched_vortragslisten:
                    issues.append("Vortragsliste fehlt.")
                else:
                    attachments.extend(matched_vortragslisten)

            sponsors.append(
                EventEndSponsorPlan(
                    row_number=sponsor.row_number,
                    sponsor_name=sponsor.sponsor_name,
                    package=SUPPORTED_PACKAGES[package_key],
                    language=sponsor.language,
                    to_email=sponsor.to_email,
                    cc_email=sponsor.cc_email,
                    contact_first_name=sponsor.contact_first_name,
                    contact_last_name=sponsor.contact_last_name,
                    attachments=tuple(attachments),
                    status="ready" if not issues else "blocked",
                    details=" ".join(issues).strip(),
                )
            )

        ready_count = sum(1 for sponsor in sponsors if sponsor.is_ready)
        blocked_count = sum(1 for sponsor in sponsors if not sponsor.is_ready)
        return EventEndAssemblyResult(
            sponsors=tuple(sponsors),
            ready_count=ready_count,
            blocked_count=blocked_count,
            skipped_count=max(candidate_rows - len(sponsors), 0),
            kontaktliste_protected=kontaktliste_protected,
            kontaktliste_protection_details=kontaktliste_protection_details,
        )
    finally:
        workbook.close()


def build_eventende_serienmails(
    sponsors: tuple[EventEndSponsorPlan, ...],
    *,
    subject_template: str,
    body_html_template: str,
    sender_email: str,
) -> list[SerienMail]:
    mails: list[SerienMail] = []
    for sponsor in sponsors:
        if not sponsor.is_ready:
            continue
        vorname = sponsor.contact_first_name or sponsor.sponsor_name
        mails.append(
            SerienMail(
                to_email=sponsor.to_email,
                cc_email=sponsor.cc_email,
                vorname=vorname,
                firma=sponsor.sponsor_name,
                subject=build_subject(
                    subject_template,
                    vorname,
                    sponsor.sponsor_name,
                    sponsor.to_email,
                ),
                html_body=render_final_mail_html(
                    body_html_template,
                    sender_email=sender_email.strip(),
                    vorname=vorname,
                    firma=sponsor.sponsor_name,
                    email=sponsor.to_email,
                ),
                attachments=sponsor.attachments,
            )
        )
    return mails


def build_eventende_summary_dataframe(
    sponsors: tuple[EventEndSponsorPlan, ...],
) -> pd.DataFrame:
    rows = [
        {
            "Sponsor": sponsor.sponsor_name,
            "Paket": sponsor.package,
            "E-Mail": sponsor.to_email,
            "Kopie": sponsor.cc_email or "-",
            "Anh\u00e4nge": " | ".join(sponsor.attachment_names) or "-",
            "Status": "Bereit" if sponsor.is_ready else "Blockiert",
            "Hinweis": sponsor.details or "-",
        }
        for sponsor in sponsors
    ]
    return pd.DataFrame(
        rows,
        columns=["Sponsor", "Paket", "E-Mail", "Kopie", "Anh\u00e4nge", "Status", "Hinweis"],
    )
