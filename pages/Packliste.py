"""
Packliste – Streamlit page
==========================

Single authoritative truth: st.session_state["pl_df"]

Architecture (optimised for fluid checkbox UX)
───────────────────────────────────────────────
  • Checkbox click  → @st.fragment reruns (fast, ~50-100 ms)
                    → NO full-app rerun triggered per click
                    → NO widget remount, NO scroll reset
  • on_change callback fires BEFORE the fragment body re-executes:
      _sync_edits / _sync_verladen
        → pl_df updated
        → _auto_save()  (sidecar sync + async Excel)
        → pl_version NOT bumped  ← key stays stable, widget stays alive
  • Fragment body runs with the SAME pl_version:
      → SAME key → SAME widget instance → SAME frozen_base
      → edited_rows preserved intact by Streamlit
      → visual = frozen_base + edited_rows = pl_df  ✓

State layers
────────────
  pl_df         sole authoritative truth; only _sync_edits/_sync_verladen write here
  frozen_base   stable view-basis for st.data_editor; NEVER modified between
                undo/load events; created once per (tab, session/version)
  edited_rows   Streamlit's in-flight delta from frozen_base; accumulates across
                all clicks on the same widget instance; never discarded as long as
                the base and key stay the same
  pl_version    bumped ONLY on undo and file-load (intentional full rebuilds)

Why the visual stays correct without per-click version bumps:
  The editor renders frozen_base + edited_rows.  After each commit:
    frozen_base[row] = initial value (unchanged)
    edited_rows[row] = user's latest value
    → visual = user's latest value = pl_df value  ✓
  As long as the key and base are stable, Streamlit preserves edited_rows
  across fragment reruns, so every accumulated click remains visible.
  There is no race condition because there is no full-app rerun per click —
  the only scenario where edited_rows could be lost by Streamlit is when the
  base DataFrame changes or a new key is used, neither of which happens here.

Why this did NOT regress the old "UI empty, model correct" bug:
  That bug required a full-app rerun (which could reset edited_rows via a
  stale browser delta).  With no full-app rerun per click, that race cannot
  occur.  On browser-refresh the sidecar is loaded, pl_version resets to 1,
  and a fresh frozen_base is built from the restored pl_df — both start from
  the same committed state, so the visual is immediately correct.

Global metrics live in the separate "Übersicht" tab, decoupled from the
fast click path.

Persistence (two-stage save)
─────────────────────────────
  Stage 1 – Sidecar (sync, ~1 ms, atomic pickle next to the Excel file)
      Written synchronously inside the on_change callback.
      Data survives any browser-refresh from this point on.
      File: <excel_stem>_autosave.pkl

  Stage 2 – Excel (async background thread, generation-tracked)
      Only the LATEST generation snapshot reaches the file.
      Near-atomic: writes to <stem>.tmp.xlsx first, then renames.
      No st.* calls inside the worker thread.

Startup / browser-refresh
──────────────────────────
  1. Read .packliste_meta.json → last-active Excel path.
  2. restore_state_for_path(path):
       sidecar newer than Excel → load from sidecar + kick off Excel sync.
       otherwise               → load from Excel directly.
  3. Session-state populated, page renders – user sees previous state instantly.

Row identity
────────────
  Packen/Ersetzen/Definitionen: frozen_base.index[pos] is the original pandas
  integer index (0-based from pd.read_excel), preserved through sort/filter.
  Maps 1-to-1 to pl_df rows regardless of how the view is filtered/sorted.
  Verladen: content-based matching (Bereich + Kategorie) via groupby aggregation.
  Both strategies are stable; no separate row-ID column is needed.
"""

import json
import logging
import pickle
import threading
import time
from io import BytesIO
from pathlib import Path

import openpyxl
import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from streamlit_ui import apply_page_title_style, page_title_html

st.set_page_config(page_title="Packliste", layout="wide")

# ── Logging ────────────────────────────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(name)s %(message)s",
)
_log = logging.getLogger("packliste")

# ── Constants ──────────────────────────────────────────────────────────────────

_ERROR_TOAST_CSS = """<style>
div[data-testid="stToast"] {
    background-color: #F1C40F !important;
    border: none !important;
}
div[data-testid="stToast"] p,
div[data-testid="stToast"] span,
div[data-testid="stToast"] div {
    color: #000000 !important;
}
</style>"""

_CRITICAL_TOAST_CSS = """<style>
div[data-testid="stToast"] {
    background-color: #C0392B !important;
    border: none !important;
}
div[data-testid="stToast"] p,
div[data-testid="stToast"] span,
div[data-testid="stToast"] div {
    color: white !important;
}
</style>"""

_SUCCESS_TOAST_CSS = """<style>
div[data-testid="stToast"] {
    background-color: #27AE60 !important;
    border: none !important;
}
div[data-testid="stToast"] p,
div[data-testid="stToast"] span,
div[data-testid="stToast"] div {
    color: white !important;
}
</style>"""

SHEET_NAME        = "Packliste"
HEADER_ROW_IDX    = 2
DATA_START_XL_ROW = 4
CHECKED           = "☒"
UNCHECKED         = "☐"
CHECKBOX_COLS     = [
    "Verpackt", "Nachfüllen", "Reinigung",
    "kurz vor Event packen", "Verladen",
]
DEFAULT_PATH = r"C:\Users\admin\Downloads\26BER_Packliste.xlsx"
MAX_HISTORY  = 30

REQUIRED_COLS = [
    "Bereich", "Gegenstand", "Beschreibung", "Menge",
    "Kategorie", "Notizen", "Verpackt", "Nachfüllen",
    "Reinigung", "kurz vor Event packen",
]

# Persists last-active Excel path across browser refreshes
_META_FILE = Path(__file__).parent / ".packliste_meta.json"

# ── Thread-shared state (persisted via cache_resource across reruns) ──────────
#
# Streamlit re-executes the entire script on every rerun.  A plain module-level
# assignment like `_excel_result = {...}` would therefore RESET the dict to
# {"status": "ok"} on every rerun — the background thread's "error" result
# would be lost immediately and the save button would always show "Gespeichert".
# The same reset would zero out _save_gen, causing the gen-check inside the
# worker to fail (thread gen ≠ 0) and skip the write entirely.
#
# st.cache_resource returns the SAME object on every rerun, so the background
# thread and the current script run always share the identical dict / list /
# Lock instances.

@st.cache_resource
def _get_thread_state() -> dict:
    return {
        "save_lock":     threading.Lock(),
        "save_gen_lock": threading.Lock(),
        "save_gen":      [0],
        "excel_result":  {"status": "ok", "error": None},
    }

_ts            = _get_thread_state()
_save_lock     = _ts["save_lock"]
_save_gen_lock = _ts["save_gen_lock"]
_save_gen      = _ts["save_gen"]
_excel_result  = _ts["excel_result"]


# ── Meta-file helpers (last-active path) ──────────────────────────────────────

def save_last_active_path(path: str) -> None:
    """Persist the last-used Excel path so refreshes can restore it."""
    try:
        with open(_META_FILE, "w", encoding="utf-8") as fh:
            json.dump({"path": path, "ts": time.time()}, fh)
    except Exception as exc:
        _log.warning("Cannot write meta %s: %s", _META_FILE, exc)


def load_last_active_path() -> str:
    """Return last-active path, or DEFAULT_PATH when unavailable."""
    try:
        if _META_FILE.exists():
            with open(_META_FILE, "r", encoding="utf-8") as fh:
                data = json.load(fh)
            p = data.get("path", "")
            if p and Path(p).exists():
                return p
    except Exception as exc:
        _log.warning("Cannot read meta %s: %s", _META_FILE, exc)
    return DEFAULT_PATH


# ── Sidecar helpers ────────────────────────────────────────────────────────────

def _sidecar_path(excel_path: str) -> Path:
    p = Path(excel_path)
    return p.parent / f"{p.stem}_autosave.pkl"


def _write_sidecar(df: pd.DataFrame, excel_path: str) -> None:
    """Atomic synchronous pickle write.  Raises on failure – caller handles."""
    sp      = _sidecar_path(excel_path)
    tmp     = sp.with_suffix(".pkl.tmp")
    payload = {"df": df.copy(), "path": excel_path, "ts": time.time()}
    with open(tmp, "wb") as fh:
        pickle.dump(payload, fh, protocol=pickle.HIGHEST_PROTOCOL)
    tmp.replace(sp)


def _read_sidecar(excel_path: str):
    sp = _sidecar_path(excel_path)
    if not sp.exists():
        return None
    try:
        with open(sp, "rb") as fh:
            data = pickle.load(fh)
        if data.get("path") != excel_path:
            return None
        return data["df"]
    except Exception as exc:
        _log.warning("Cannot read sidecar %s: %s", sp, exc)
        return None


def _sidecar_newer(excel_path: str) -> bool:
    sp = _sidecar_path(excel_path)
    ep = Path(excel_path)
    if not sp.exists():
        return False
    if not ep.exists():
        return True
    return sp.stat().st_mtime > ep.stat().st_mtime


# ── State restoration ──────────────────────────────────────────────────────────

def restore_state_for_path(path: str):
    """
    Load best available state for *path*.
    Returns (df, source) where source ∈ {"sidecar", "excel"}.
    Raises ValueError / IOError on failure.
    """
    if _sidecar_newer(path):
        sdf = _read_sidecar(path)
        if sdf is not None:
            return sdf, "sidecar"
    df = load_df(path)
    return df, "excel"


# ── Excel I/O ──────────────────────────────────────────────────────────────────

def _detect_sheet_layout(ws) -> tuple[int, int, dict[str, int]]:
    """Return header row index for pandas, first data row in Excel, and header map."""
    for excel_row in range(1, min(ws.max_row, 12) + 1):
        col_map = {
            str(cell.value).strip(): cell.column
            for cell in ws[excel_row]
            if cell.value and str(cell.value).strip()
        }
        if all(col in col_map for col in REQUIRED_COLS):
            data_start_row = excel_row + 1
            for candidate_row in range(excel_row + 1, ws.max_row + 1):
                has_data = any(
                    ws.cell(candidate_row, xl_col).value not in (None, "")
                    for xl_col in col_map.values()
                )
                if has_data:
                    data_start_row = candidate_row
                    break
            return excel_row - 1, data_start_row, col_map

    raise ValueError(
        "Headerzeile im Sheet 'Packliste' nicht gefunden. "
        "Bitte prÃ¼fen, ob die Pflichtspalten vorhanden sind."
    )


def load_df(path: str) -> pd.DataFrame:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
    try:
        ws = wb[SHEET_NAME]
        header_row_idx, _, _ = _detect_sheet_layout(ws)
    finally:
        wb.close()

    df = pd.read_excel(path, sheet_name=SHEET_NAME,
                       header=header_row_idx, dtype=str)
    df = df.where(df.notna() & (df != "nan"), "")
    missing = [c for c in REQUIRED_COLS if c not in df.columns]
    if missing:
        raise ValueError(
            f"Fehlende Spalten: {missing}\n\n"
            f"Gefundene Spalten: {list(df.columns)}\n\n"
            "Bitte prüfen, ob die richtige Datei und das Sheet 'Packliste' geladen wird."
        )
    for col in CHECKBOX_COLS:
        if col in df.columns:
            df[col] = df[col] == CHECKED
    if "Verladen" not in df.columns:
        df["Verladen"] = False
    return df


def save_df(df: pd.DataFrame, path: str) -> None:
    """
    Near-atomic Excel write:
      1. Write to <stem>.tmp.xlsx
      2. Rename to <path>  (replaces original only after full successful write)
    If anything fails the original file is untouched; tmp is cleaned up.
    """
    p   = Path(path)
    tmp = p.parent / (p.stem + ".tmp.xlsx")
    try:
        wb   = openpyxl.load_workbook(path)
        ws   = wb[SHEET_NAME]
        header_row_idx, data_start_row, col_map = _detect_sheet_layout(ws)
        hrow = header_row_idx + 1
        if "Verladen" not in col_map:
            nxt = max(col_map.values()) + 1
            ws.cell(row=hrow, column=nxt, value="Verladen")
            col_map["Verladen"] = nxt
        for i, (_, row) in enumerate(df.iterrows()):
            xl_row = data_start_row + i
            for col_name, xl_col in col_map.items():
                if col_name not in df.columns:
                    continue
                val = row[col_name]
                if col_name in CHECKBOX_COLS:
                    val = CHECKED if bool(val) else UNCHECKED
                else:
                    val = val if val != "" else None
                ws.cell(row=xl_row, column=xl_col, value=val)
        wb.save(str(tmp))
        tmp.replace(p)
    except Exception:
        try:
            tmp.unlink(missing_ok=True)
        except Exception:
            pass
        raise


# ── Column-config shortcuts ────────────────────────────────────────────────────

def _cc_text(label: str, width: str = "medium") -> st.column_config.TextColumn:
    return st.column_config.TextColumn(label, width=width)

def _cc_number_left(label: str, width: str | None = None, fmt: str = "%d") -> dict:
    cfg = st.column_config.NumberColumn(label, width=width, format=fmt)
    cfg["alignment"] = "left"
    return cfg

def _cc_check(label: str) -> st.column_config.CheckboxColumn:
    return st.column_config.CheckboxColumn(label)


def _to_excel_bytes(df: pd.DataFrame, sheet_name: str = "Liste") -> bytes:
    data = BytesIO()
    export_df = df.reset_index(drop=True).copy()
    for col in CHECKBOX_COLS:
        if col in export_df.columns:
            export_df[col] = export_df[col].map(lambda v: "Ja" if bool(v) else "Nein")
    with pd.ExcelWriter(data, engine="openpyxl") as writer:
        export_df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.book[sheet_name]

        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style="thin", color="D9E2F3"),
            right=Side(style="thin", color="D9E2F3"),
            top=Side(style="thin", color="D9E2F3"),
            bottom=Side(style="thin", color="D9E2F3"),
        )
        stripe_fill = PatternFill("solid", fgColor="F7F9FC")

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        ws.sheet_view.showGridLines = False
        ws.row_dimensions[1].height = 24

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border

        for row_idx in range(2, ws.max_row + 1):
            is_striped = row_idx % 2 == 0
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = border
                cell.alignment = cell_alignment
                if is_striped:
                    cell.fill = stripe_fill

        for col_idx, column_name in enumerate(export_df.columns, start=1):
            col_letter = get_column_letter(col_idx)
            values = [column_name, *export_df.iloc[:, col_idx - 1].tolist()]
            max_length = max(len(str(v)) if v is not None else 0 for v in values)
            width = min(max(max_length + 2, 12), 42)
            if column_name in {"Beschreibung", "Notizen"}:
                width = min(max(max_length + 2, 24), 56)
            if column_name in CHECKBOX_COLS or column_name == "Fortschritt":
                width = min(max(max_length + 2, 12), 16)
                for row_idx in range(2, ws.max_row + 1):
                    ws.cell(row=row_idx, column=col_idx).alignment = center_alignment
            ws.column_dimensions[col_letter].width = width

        if "Fortschritt" in export_df.columns:
            progress_idx = export_df.columns.get_loc("Fortschritt") + 1
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=progress_idx).number_format = '0" %"'
    data.seek(0)
    return data.getvalue()


def _download_filename(slug: str) -> str:
    path = Path(st.session_state.get("pl_path") or "packliste.xlsx")
    return f"{path.stem}_{slug}.xlsx"


def _render_excel_download(
    df: pd.DataFrame,
    *,
    slug: str,
    key: str,
    label: str = "Liste als Excel herunterladen",
    sheet_name: str = "Liste",
) -> None:
    st.download_button(
        label,
        data=_to_excel_bytes(df, sheet_name=sheet_name),
        file_name=_download_filename(slug),
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=key,
    )


def _build_overview_stats(
    df: pd.DataFrame,
    *,
    done_mask: pd.Series,
    total_mask: pd.Series | None = None,
    value_label: str,
) -> pd.DataFrame:
    if "Bereich" not in df.columns:
        return pd.DataFrame(columns=["Bereich", value_label, "Gesamt", "Fortschritt"])

    if total_mask is None:
        total_mask = pd.Series(True, index=df.index)

    base = df.loc[total_mask.astype(bool), ["Bereich"]].copy()
    if base.empty:
        return pd.DataFrame(columns=["Bereich", value_label, "Gesamt", "Fortschritt"])

    base[value_label] = done_mask.loc[base.index].astype(bool).astype(int)
    base["_row_count"] = 1
    stats = (
        base.groupby("Bereich", as_index=False)
        .agg({value_label: "sum", "_row_count": "sum"})
    )
    stats = stats.rename(columns={"_row_count": "Gesamt"})
    stats["Fortschritt"] = (
        stats[value_label] / stats["Gesamt"].clip(lower=1) * 100
    ).round(0).astype(int)

    total_row = pd.DataFrame(
        {
            "Bereich": ["Gesamt"],
            value_label: [int(base[value_label].sum())],
            "Gesamt": [len(base)],
        }
    )
    total_row["Fortschritt"] = (
        total_row[value_label] / total_row["Gesamt"].clip(lower=1) * 100
    ).round(0).astype(int)

    return pd.concat([stats.sort_values("Bereich"), total_row], ignore_index=True)


def _render_overview_stats_table(
    title: str,
    stats: pd.DataFrame,
    *,
    value_label: str,
    row_color: str = "#3B82F6",
    total_color: str = "#0F766E",
    note: str | None = None,
) -> None:
    _render_overview_stats_df(title, stats, value_label=value_label, note=note)
    return

    st.subheader(title)
    if note:
        st.caption(note)
    if stats.empty:
        st.info("Keine Daten verfügbar.")
        return

    rows_html: list[str] = []
    for _, row in stats.iterrows():
        bereich = str(row["Bereich"])
        value = int(row[value_label])
        total = int(row["Gesamt"])
        progress = max(0, min(int(row["Fortschritt"]), 100))
        is_total = bereich == "Gesamt"
        bar_color = total_color if is_total else row_color
        weight = "700" if is_total else "500"
        border_top = "border-top:2px solid #CBD5E1;" if is_total else ""

        rows_html.append(
            f"""
            <tr style="font-weight:{weight};{border_top}">
                <td>{html.escape(bereich)}</td>
                <td style="text-align:right;">{value}</td>
                <td style="text-align:right;">{total}</td>
                <td>
                    <div style="display:flex;align-items:center;gap:12px;">
                        <div style="flex:1;height:10px;background:#E8EDF3;border-radius:999px;overflow:hidden;">
                            <div style="width:{progress}%;height:100%;background:{bar_color};border-radius:999px;"></div>
                        </div>
                        <span style="min-width:46px;text-align:right;color:#475569;">{progress} %</span>
                    </div>
                </td>
            </tr>
            """
        )

    st.markdown(
        f"""
        <div class="ov-stats-card" style="border:1px solid #E5E7EB;border-radius:14px;overflow:hidden;margin-bottom:1rem;">
            <table class="ov-stats-table" style="width:100%;border-collapse:collapse;font-size:0.98rem;">
                <thead>
                    <tr style="background:#F8FAFC;color:#64748B;">
                        <th style="text-align:left;padding:12px 14px;border-bottom:1px solid #E5E7EB;">Bereich</th>
                        <th style="text-align:right;padding:12px 14px;border-bottom:1px solid #E5E7EB;">{html.escape(value_label)}</th>
                        <th style="text-align:right;padding:12px 14px;border-bottom:1px solid #E5E7EB;">Gesamt</th>
                        <th style="text-align:left;padding:12px 14px;border-bottom:1px solid #E5E7EB;">Fortschritt</th>
                    </tr>
                </thead>
                <tbody>
                    {"".join(rows_html)}
                </tbody>
            </table>
        </div>
        <style>
        .ov-stats-card .ov-stats-table td {{
            padding: 12px 14px;
            border-bottom: 1px solid #EEF2F7;
            background: #FFFFFF;
        }}
        .ov-stats-card .ov-stats-table tbody tr:last-child td {{
            border-bottom: none;
        }}
        </style>
        """,
        unsafe_allow_html=True,
    )


# ── History ────────────────────────────────────────────────────────────────────

def _render_overview_stats_df(
    title: str,
    stats: pd.DataFrame,
    *,
    value_label: str,
    note: str | None = None,
) -> None:
    st.subheader(title)
    if note:
        st.caption(note)
    if stats.empty:
        st.info("Keine Daten verfuegbar.")
        return

    view = stats.loc[:, ["Bereich", value_label, "Gesamt", "Fortschritt"]].copy()
    view["Fortschritt"] = view["Fortschritt"].astype(int).clip(lower=0, upper=100)
    progress_cfg = st.column_config.ProgressColumn(
        "Fortschritt",
        format="%d%%",
        min_value=0,
        max_value=100,
    )
    progress_cfg["alignment"] = "left"

    st.dataframe(
        view,
        use_container_width=True,
        hide_index=True,
        column_config={
            "Bereich": st.column_config.TextColumn("Bereich", width="medium"),
            value_label: _cc_number_left(value_label, fmt="%d"),
            "Gesamt": _cc_number_left("Gesamt", fmt="%d"),
            "Fortschritt": progress_cfg,
        },
    )


def _push_history() -> None:
    st.session_state["pl_redo_stack"] = []
    hist: list = st.session_state.setdefault("pl_history", [])
    hist.append(st.session_state["pl_df"].copy())
    if len(hist) > MAX_HISTORY:
        hist.pop(0)


# ── Two-stage save pipeline ────────────────────────────────────────────────────

def _auto_save() -> None:
    """
    Stage 1 – Sidecar (sync, ~1 ms):
        Atomic pickle.  Data survives any refresh from this point on.
        Errors surfaced immediately via st.toast (no silent swallowing).

    Stage 2 – Excel (async background thread, generation-tracked):
        Only the latest generation writes; older snapshots are dropped.
        No st.* calls inside the worker.
    """
    df_snapshot = st.session_state["pl_df"].copy()
    path        = st.session_state["pl_path"]

    # Stage 1: synchronous sidecar ────────────────────────────────────────────
    try:
        _write_sidecar(df_snapshot, path)
        st.session_state.pop("pl_sidecar_err", None)
    except Exception as exc:
        msg = str(exc)
        _log.error("Sidecar write failed: %s", msg)
        st.session_state["pl_sidecar_err"] = msg
        st.toast(f"⚠️ Autosave-Fehler – Refresh könnte Daten verlieren: {msg}", icon="⚠️")

    # Stage 2: async Excel ────────────────────────────────────────────────────
    with _save_gen_lock:
        _save_gen[0] += 1
        my_gen = _save_gen[0]
    _excel_result["status"] = "saving"
    _excel_result["error"]  = None

    def _worker(df_copy: pd.DataFrame, p: str, gen: int) -> None:
        time.sleep(1.0)   # batch window – coalesces rapid edits
        with _save_gen_lock:
            if gen != _save_gen[0]:
                _log.info("Excel save gen %d superseded by %d – skipped",
                          gen, _save_gen[0])
                return
        with _save_lock:
            with _save_gen_lock:
                if gen != _save_gen[0]:
                    return
            try:
                save_df(df_copy, p)
                _excel_result["status"] = "ok"
                _excel_result["error"]  = None
                _log.info("Excel save gen %d complete", gen)
            except Exception as exc:
                _excel_result["status"] = "error"
                _excel_result["error"]  = str(exc)
                _log.error("Excel save gen %d failed: %s", gen, exc)

    threading.Thread(
        target=_worker, args=(df_snapshot, path, my_gen), daemon=True
    ).start()


# ── Frozen-base / edit-sync helpers ───────────────────────────────────────────
#
# Frozen-base invariant:
#   The DataFrame passed to st.data_editor is NEVER modified after creation.
#   Streamlit accumulates edited_rows as a delta on the base it last received.
#   If the base changes between renders, Streamlit discards all accumulated
#   edited_rows — causing the "second click disappears" bug.  Keeping the
#   base frozen (and the widget key stable) prevents any such reset.
#
# Why we do NOT bump pl_version on every commit:
#   A version bump → new key → Streamlit mounts a NEW widget → DOM remount
#   → scroll position lost → visible table flicker on every click.
#   Instead the same widget instance is reused across all clicks.
#   The visual stays correct because edited_rows accumulates the full delta
#   from the frozen_base and is never discarded (stable key + stable base).
#   pl_version is bumped ONLY on intentional full-rebuilds (undo, file-load).

def _get_frozen_base(key: str, initial: pd.DataFrame) -> pd.DataFrame:
    """
    Return the frozen base for *key*, creating it from *initial* if absent.

    *initial* is always passed as a view derived from CURRENT pl_df, so a
    fresh base starts visually identical to pl_df.

    On new version key: stale bases with the same prefix are evicted so
    session_state doesn't grow indefinitely.
    """
    if key not in st.session_state:
        prefix = key.rsplit("_", 1)[0] + "_"
        stale  = [k for k in st.session_state if k.startswith(prefix) and k != key]
        for k in stale:
            _log.debug("Evicting stale frozen base: %s", k)
            del st.session_state[k]
        st.session_state[key] = initial.copy()
        _log.debug("Created frozen base %s (%d rows)", key, len(initial))
    return st.session_state[key]


def _sync_edits(editor_key: str, frozen_base: pd.DataFrame) -> None:
    """
    Apply new differences from edited_rows into pl_df.
    pl_version is NOT bumped — the same widget key and frozen_base are reused
    on the next fragment run so the editor stays alive without a remount.
    """
    state = st.session_state.get(editor_key)
    if not isinstance(state, dict):
        return
    edited_rows: dict = state.get("edited_rows", {})
    if not edited_rows:
        return

    df             = st.session_state["pl_df"]
    history_pushed = False
    changed_cells: list = []

    for pos_key, changes in edited_rows.items():
        pos = int(pos_key)
        if pos >= len(frozen_base):
            continue
        df_idx = frozen_base.index[pos]
        for col, new_val in changes.items():
            if col not in df.columns:
                continue
            old_val = df.at[df_idx, col]
            if old_val != new_val:
                if not history_pushed:
                    _push_history()
                    history_pushed = True
                df.at[df_idx, col] = new_val
                changed_cells.append((df_idx, col, old_val, new_val))

    if history_pushed:
        _log.info(
            "Commit %s: ver=%d, cells=%s",
            editor_key, st.session_state["pl_version"],
            [(idx, c, f"{ov!r}→{nv!r}") for idx, c, ov, nv in changed_cells],
        )
        _auto_save()
        # Kein st.rerun() hier — on_change-Callbacks ignorieren Rerun (No-Op).
        # Fragment ruft _flush_full_rerun_after_editor_commit() auf.
        st.session_state["_pl_rerun_after_commit"] = True


def _sync_verladen(editor_key: str, frozen_base: pd.DataFrame) -> None:
    """
    Like _sync_edits but propagates Verladen to every row sharing the same
    Bereich + Kategorie.  pl_version is NOT bumped — widget stays alive.
    """
    state = st.session_state.get(editor_key)
    if not isinstance(state, dict):
        return
    edited_rows: dict = state.get("edited_rows", {})
    if not edited_rows:
        return

    df             = st.session_state["pl_df"]
    history_pushed = False
    changed_groups: list = []

    for pos_key, changes in edited_rows.items():
        pos = int(pos_key)
        if pos >= len(frozen_base):
            continue
        bereich   = frozen_base.iloc[pos]["Bereich"]
        kategorie = frozen_base.iloc[pos]["Kategorie"]
        new_val   = bool(changes.get("Verladen", frozen_base.iloc[pos]["Verladen"]))
        mask      = (df["Bereich"] == bereich) & (df["Kategorie"] == kategorie)
        if not (df.loc[mask, "Verladen"] == new_val).all():
            if not history_pushed:
                _push_history()
                history_pushed = True
            df.loc[mask, "Verladen"] = new_val
            changed_groups.append((bereich, kategorie, new_val))

    if history_pushed:
        _log.info(
            "Commit %s: ver=%d, verladen groups=%s",
            editor_key, st.session_state["pl_version"],
            [(b, k, nv) for b, k, nv in changed_groups],
        )
        _auto_save()
        st.session_state["_pl_rerun_after_commit"] = True


def _flush_full_rerun_after_editor_commit() -> None:
    """Nach Editor-Commit: voller Rerun, damit Header (Undo) pl_history sieht."""
    if st.session_state.pop("_pl_rerun_after_commit", False):
        st.rerun()


# ── Callback factory ───────────────────────────────────────────────────────────

def _make_callback(editor_key: str, frozen_base: pd.DataFrame,
                   verladen: bool = False):
    """
    Returns an on_change callback for a data_editor.

    Streamlit calls this callback BEFORE the fragment body re-executes.
    The callback commits changes to pl_df and triggers autosave.
    pl_version is NOT bumped → same key → same widget instance is reused →
    no remount, no scroll reset.

    The frozen_base reference is captured at creation time.  Since it is
    never modified (frozen-base invariant), it stays valid for all future
    invocations of this callback.
    """
    def _cb() -> None:
        if verladen:
            _sync_verladen(editor_key, frozen_base)
        else:
            _sync_edits(editor_key, frozen_base)
    return _cb


# ── Session-state init ─────────────────────────────────────────────────────────

for _k, _v in [("pl_df", None), ("pl_path", DEFAULT_PATH),
               ("pl_version", 0), ("pl_history", []), ("pl_redo_stack", [])]:
    if _k not in st.session_state:
        st.session_state[_k] = _v


# ── Auto-load on startup / browser-refresh ─────────────────────────────────────
# st.session_state is empty after a browser-refresh → restore transparently.

if st.session_state["pl_df"] is None:
    _path = load_last_active_path()
    try:
        _df, _src = restore_state_for_path(_path)
        st.session_state["pl_df"]      = _df
        st.session_state["pl_path"]    = _path
        st.session_state["pl_version"] = 1
        save_last_active_path(_path)
        if _src == "sidecar":
            st.session_state["_al_msg"] = "🔄 Stand aus Autosave wiederhergestellt."
            _auto_save()   # sync sidecar state back to Excel in background
    except Exception as _exc:
        st.session_state["_al_err"] = str(_exc)


# ── Header ─────────────────────────────────────────────────────────────────────
# Minimal: title + save status + undo.
# Global metrics are in the "Übersicht" tab → not part of the click flow.

if _excel_result["status"] == "error":
    st.session_state["pl_excel_err"] = _excel_result["error"]
elif _excel_result["status"] == "ok":
    st.session_state.pop("pl_excel_err", None)

apply_page_title_style()

hist_len = len(st.session_state.get("pl_history", []))
redo_len = len(st.session_state.get("pl_redo_stack", []))
title_col, btn_col = st.columns([7, 2])

with title_col:
    st.markdown(page_title_html("Packliste"), unsafe_allow_html=True)

with btn_col:
    st.markdown("<div style='margin-top:1.1rem'></div>", unsafe_allow_html=True)
    save_col, undo_col, redo_col = st.columns(3)
    with save_col:
        if st.button("💾", use_container_width=True, key="save_status_btn",
                     help="Speichern"):
            # Neuen Speicherversuch starten (Retry), damit ein zuvor fehlgeschlagener
            # Save nach Schließen der Excel-Datei erneut versucht wird.
            if st.session_state.get("pl_df") is not None:
                _auto_save()
            time.sleep(1.8)   # warten bis Hintergrund-Thread Schreibversuch abgeschlossen hat
            # _excel_result direkt lesen (nach sleep aktuell vom Thread befüllt)
            if st.session_state.get("pl_sidecar_err"):
                st.markdown(_CRITICAL_TOAST_CSS, unsafe_allow_html=True)
                st.toast("Autosave-Fehler!")
            elif _excel_result["status"] == "error":
                st.session_state["pl_excel_err"] = _excel_result["error"]
                st.markdown(_ERROR_TOAST_CSS, unsafe_allow_html=True)
                st.toast("Excel nicht aktualisiert – Excel-Datei schließen und erneut speichern")
            elif _excel_result["status"] == "saving":
                st.toast("⏳ Speichern läuft noch…")
            else:
                st.session_state.pop("pl_excel_err", None)
                st.markdown(_SUCCESS_TOAST_CSS, unsafe_allow_html=True)
                st.toast("Gespeichert")
    with undo_col:
        undo_label = "↩"
        if st.button(undo_label, disabled=(hist_len == 0),
                     use_container_width=True, key="undo_btn",
                     help="Letzte Änderung rückgängig machen"):
            redo = st.session_state.setdefault("pl_redo_stack", [])
            redo.append(st.session_state["pl_df"].copy())
            if len(redo) > MAX_HISTORY:
                redo.pop(0)
            prev = st.session_state["pl_history"].pop()
            st.session_state["pl_df"]      = prev
            st.session_state["pl_version"] += 1
            _auto_save()
            st.rerun()
    with redo_col:
        if st.button("↪", disabled=(redo_len == 0),
                     use_container_width=True, key="redo_btn",
                     help="Wiederherstellen"):
            hist = st.session_state.setdefault("pl_history", [])
            hist.append(st.session_state["pl_df"].copy())
            if len(hist) > MAX_HISTORY:
                hist.pop(0)
            st.session_state["pl_df"] = st.session_state["pl_redo_stack"].pop()
            st.session_state["pl_version"] += 1
            _auto_save()
            st.rerun()

# Fehler-Meldungen unterhalb des Headers, volle Breite
if st.session_state.get("pl_sidecar_err"):
    st.error(f"⚠️ Autosave-Fehler: {st.session_state['pl_sidecar_err']}")
elif st.session_state.get("pl_excel_err"):
    st.warning(
        f"⚠️ Excel nicht aktualisiert (Daten in Autosave gesichert): "
        f"{st.session_state['pl_excel_err']}"
    )
if msg := st.session_state.pop("_al_msg", None):
    st.info(msg)
if err := st.session_state.pop("_al_err", None):
    st.error(f"Fehler beim Auto-Laden: {err}")


# ── File loader ────────────────────────────────────────────────────────────────

with st.expander(
    "📂 Datei laden" if st.session_state["pl_df"] is None else "📂 Andere Datei laden",
    expanded=st.session_state["pl_df"] is None,
):
    path_input = st.text_input(
        "Dateipfad zur Excel-Datei",
        value=st.session_state["pl_path"],
        key="pl_path_input",
    )
    if st.button("Laden", key="load_btn", type="primary"):
        try:
            new_df = load_df(path_input)
            _write_sidecar(new_df, path_input)
            save_last_active_path(path_input)
            st.session_state["pl_df"]      = new_df
            st.session_state["pl_path"]    = path_input
            st.session_state["pl_version"] += 1
            st.session_state["pl_history"] = []
            st.session_state["pl_redo_stack"] = []
            st.rerun()   # full-app rerun intentional: resets all tab views
        except Exception as exc:
            st.error(f"Fehler beim Laden:\n{exc}")

if st.session_state["pl_df"] is None:
    st.stop()

# Achtung, wenn Gebrauchsgegenstände (Ersetzen) Notizen haben
df = st.session_state["pl_df"]
if "Nachfüllen" in df.columns and "Notizen" in df.columns:
    ers_df = df.loc[~df["Nachfüllen"].astype(bool), "Notizen"]
    hat_notizen = ers_df.apply(
        lambda v: bool(pd.notna(v) and str(v).strip())
    ).any()
    if hat_notizen:
        st.warning(
            "⚠️ Achtung: Es gibt offene Notizen bei den Gebrauchsgegenständen "
            "(Tab Ersetzen)."
        )

st.divider()

# ── Tabs ───────────────────────────────────────────────────────────────────────

tab_pack, tab_verl, tab_ers, tab_ueb, tab_def = st.tabs(
    ["📦 Packen", "🚚 Verladen", "🔄 Ersetzen", "📊 Übersicht", "📋 Definitionen"]
)


# ── Fragments ──────────────────────────────────────────────────────────────────
# Each fragment is self-contained: it reads pl_df, builds its frozen_base,
# renders its editor(s), and registers its on_change callback.
# Nach Commit setzen _sync_* ein Flag; Fragment startet mit
# _flush_full_rerun_after_editor_commit() (Rerun nur außerhalb Callback).

@st.fragment
def _frag_packen() -> None:
    _flush_full_rerun_after_editor_commit()
    ver  = st.session_state["pl_version"]
    df   = st.session_state["pl_df"]
    pcfg = {
        "Bereich":    _cc_text("Bereich",    width="small"),
        "Gegenstand": _cc_text("Gegenstand", width="medium"),
        "Menge":      _cc_text("Menge",      width="small"),
        "Kategorie":  _cc_text("Kategorie",  width="small"),
        "Verpackt":   _cc_check("Verpackt"),
    }

    # ── Jetzt reinigen und packen ──────────────────────────────────────────────
    st.subheader("Jetzt reinigen und packen")
    cr     = ["Bereich", "Gegenstand", "Menge", "Kategorie", "Verpackt"]
    bk_r   = f"b_rein_{ver}"
    view_r = df.loc[df["Reinigung"].astype(bool), cr].sort_values("Bereich")
    base_r = _get_frozen_base(
        bk_r, view_r
    )
    ek_r = f"e_rein_{ver}"
    if base_r.empty:
        st.info("Keine Positionen mit angekreuztem Reinigung-Kästchen.")
    else:
        st.caption(f"{len(base_r)} Position(en)")
        st.data_editor(
            base_r, key=ek_r, use_container_width=True,
            hide_index=True, num_rows="fixed",
            on_change=_make_callback(ek_r, base_r),
            column_config={k: v for k, v in pcfg.items() if k in cr},
        )
    _render_excel_download(
        view_r,
        slug="jetzt_reinigen_und_packen",
        key=f"dl_rein_{ver}",
        sheet_name="Jetzt reinigen",
    )

    st.divider()

    # ── Jetzt packen ──────────────────────────────────────────────────────────
    st.subheader("Jetzt packen – Verbrauchsgegenstände")
    cn     = ["Bereich", "Gegenstand", "Menge", "Kategorie", "Verpackt"]
    bk_n   = f"b_nach_{ver}"
    view_n = df.loc[df["Nachfüllen"].astype(bool), cn].sort_values("Bereich")
    base_n = _get_frozen_base(
        bk_n, view_n
    )
    ek_n = f"e_nach_{ver}"
    if base_n.empty:
        st.info("Keine Positionen mit angekreuztem Nachfüllen-Kästchen.")
    else:
        st.caption(f"{len(base_n)} Position(en)")
        st.data_editor(
            base_n, key=ek_n, use_container_width=True,
            hide_index=True, num_rows="fixed",
            on_change=_make_callback(ek_n, base_n),
            column_config={k: v for k, v in pcfg.items() if k in cn},
        )
    _render_excel_download(
        view_n,
        slug="jetzt_packen_verbrauchsgegenstaende",
        key=f"dl_nach_{ver}",
        sheet_name="Jetzt packen",
    )

    st.divider()

    # ── Kurz vor Event packen ──────────────────────────────────────────────────
    st.subheader("Kurz vor Event packen")
    ce     = ["Bereich", "Gegenstand", "Menge", "Kategorie"]
    bk_e   = f"b_event_{ver}"
    view_e = df.loc[df["kurz vor Event packen"].astype(bool), ce].sort_values("Bereich")
    base_e = _get_frozen_base(
        bk_e, view_e
    )
    ek_e = f"e_event_{ver}"
    if base_e.empty:
        st.info("Keine Positionen mit angekreuztem 'Kurz vor Event packen'-Kästchen.")
    else:
        st.caption(f"{len(base_e)} Position(en)")
        st.data_editor(
            base_e, key=ek_e, use_container_width=True,
            hide_index=True, num_rows="fixed",
            on_change=_make_callback(ek_e, base_e),
            column_config={k: v for k, v in pcfg.items() if k in ce},
        )
    _render_excel_download(
        view_e,
        slug="kurz_vor_event_packen",
        key=f"dl_event_{ver}",
        sheet_name="Kurz vor Event",
    )


@st.fragment
def _frag_verladen() -> None:
    _flush_full_rerun_after_editor_commit()
    ver   = st.session_state["pl_version"]
    df    = st.session_state["pl_df"]
    view_v = (
        df.groupby(["Bereich", "Kategorie"], sort=True)["Verladen"].first().reset_index()
    )
    bk_v   = f"b_verl_{ver}"
    base_v = _get_frozen_base(
        bk_v,
        view_v,
    )
    ek_v = f"e_verl_{ver}"
    st.data_editor(
        base_v, key=ek_v, use_container_width=True,
        hide_index=True, num_rows="fixed",
        on_change=_make_callback(ek_v, base_v, verladen=True),
        column_config={
            "Bereich":   _cc_text("Bereich",   width="small"),
            "Kategorie": _cc_text("Kategorie", width="small"),
            "Verladen":  _cc_check("Verladen"),
        },
    )
    _render_excel_download(
        view_v,
        slug="verladen",
        key=f"dl_verl_{ver}",
        sheet_name="Verladen",
    )


@st.fragment
def _frag_ersetzen() -> None:
    _flush_full_rerun_after_editor_commit()
    ver      = st.session_state["pl_version"]
    df       = st.session_state["pl_df"]
    st.subheader("Gebrauchsgegenstände")
    st.caption("In die Notizen das Problem reinschreiben.")
    cols_e   = ["Bereich", "Gegenstand", "Notizen"]
    bk_ers   = f"b_ers_{ver}"
    view_ers = df.loc[~df["Nachfüllen"].astype(bool), cols_e].sort_values("Bereich")
    base_ers = _get_frozen_base(
        bk_ers, view_ers
    )
    ek_ers = f"e_ers_{ver}"
    if base_ers.empty:
        st.info("Alle Positionen haben Nachfüllen angekreuzt.")
    else:
        st.caption(f"{len(base_ers)} Position(en)")
        st.data_editor(
            base_ers, key=ek_ers, use_container_width=True,
            hide_index=True, num_rows="fixed",
            on_change=_make_callback(ek_ers, base_ers),
            column_config={
                "Bereich":    _cc_text("Bereich",    width="small"),
                "Gegenstand": _cc_text("Gegenstand", width="medium"),
                "Notizen":    _cc_text("Notizen",    width="large"),
            },
        )
    _render_excel_download(
        view_ers,
        slug="ersetzen",
        key=f"dl_ers_{ver}",
        sheet_name="Ersetzen",
    )


@st.fragment
def _frag_definitionen() -> None:
    _flush_full_rerun_after_editor_commit()
    ver  = st.session_state["pl_version"]
    df   = st.session_state["pl_df"]
    cols = ["Gegenstand", "Beschreibung", "Bereich"]
    bk   = f"b_def_{ver}"
    view_def = df[cols].sort_values("Bereich")
    base = _get_frozen_base(bk, view_def)
    ek   = f"e_def_{ver}"
    st.data_editor(
        base, key=ek, use_container_width=True,
        hide_index=True, num_rows="fixed",
        on_change=_make_callback(ek, base),
        column_config={
            "Gegenstand":   _cc_text("Gegenstand",   width="medium"),
            "Beschreibung": _cc_text("Beschreibung", width="large"),
            "Bereich":      _cc_text("Bereich",      width="small"),
        },
    )
    _render_excel_download(
        view_def,
        slug="definitionen",
        key=f"dl_def_{ver}",
        sheet_name="Definitionen",
    )


@st.fragment
def _frag_uebersicht() -> None:
    """
    Summary statistics and breakdowns, deliberately decoupled from the
    per-click fragment flow.

    Updates when:
      • The user clicks "Aktualisieren" (triggers a fragment rerun of this tab)
      • A full-app rerun occurs (undo, file load, browser refresh)

    Does NOT update on every checkbox click in the other tabs — that is
    intentional; it keeps those interactions fluid.
    """
    df   = st.session_state["pl_df"]
    path = st.session_state["pl_path"]

    hdr, path_col = st.columns([1, 5])
    with hdr:
        st.button("🔄 Aktualisieren", key="ueb_refresh")
    with path_col:
        st.caption(f"Datei: `{path}`")

    st.divider()

    # ── Top metrics ───────────────────────────────────────────────────────────
    m1, m2, m3 = st.columns(3)
    total = len(df)
    replace_col = next((col for col in df.columns if col.startswith("Nachf")), None)
    if "Verpackt" in df.columns:
        packed = int(df["Verpackt"].sum())
        m1.metric("Verpackt", f"{packed} / {total}")
    if "Notizen" in df.columns and replace_col is not None:
        replace_total = int((~df[replace_col].astype(bool)).sum())
        intact_count = int(
            (
                ~df["Notizen"].apply(lambda v: bool(pd.notna(v) and str(v).strip()))
                & ~df[replace_col].astype(bool)
            ).sum()
        )
        m2.metric("Intakt", f"{intact_count} / {replace_total}")
    if all(c in df.columns for c in ("Bereich", "Kategorie", "Verladen")):
        dedup      = df.drop_duplicates(["Bereich", "Kategorie"])
        verl_done  = int(dedup["Verladen"].sum())
        verl_total = len(dedup)
        m3.metric("Verladen", f"{verl_done} / {verl_total}")

    st.divider()

    # ── Verpackt nach Bereich ─────────────────────────────────────────────────
    if all(c in df.columns for c in ("Bereich", "Verpackt")):
        pack_stats = _build_overview_stats(
            df,
            done_mask=df["Verpackt"].astype(bool),
            value_label="Verpackt",
        )
        _render_overview_stats_df(
            "Packen",
            pack_stats,
            value_label="Verpackt",
        )

    if all(c in df.columns for c in ("Bereich", "Kategorie", "Verladen")):
        verl_df = df.drop_duplicates(["Bereich", "Kategorie"]).copy()
        verl_stats = _build_overview_stats(
            verl_df,
            done_mask=verl_df["Verladen"].astype(bool),
            value_label="Verladen",
        )
        _render_overview_stats_df(
            "Verladen",
            verl_stats,
            value_label="Verladen",
        )

    replace_col = next((col for col in df.columns if col.startswith("Nachf")), None)
    if all(c in df.columns for c in ("Bereich", "Notizen")) and replace_col is not None:
        replace_total_mask = ~df[replace_col].astype(bool)
        intact_mask = (
            ~df["Notizen"].apply(lambda v: bool(pd.notna(v) and str(v).strip()))
            & replace_total_mask
        )
        replace_stats = _build_overview_stats(
            df,
            done_mask=intact_mask,
            total_mask=replace_total_mask,
            value_label="Intakt",
        )
        _render_overview_stats_df(
            "Ersetzen",
            replace_stats,
            value_label="Intakt",
        )

    return



# ── Render ─────────────────────────────────────────────────────────────────────

with tab_pack:
    _frag_packen()
with tab_verl:
    _frag_verladen()
with tab_ers:
    _frag_ersetzen()
with tab_def:
    _frag_definitionen()
with tab_ueb:
    _frag_uebersicht()
