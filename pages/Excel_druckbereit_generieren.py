# pages/01_Excel_druckbereit_generieren.py
# Streamlit: HubSpot -> Excel (druckfertig) mit Template-Spaltenbreiten + Titel im Seitenkopf
# + Multi-Select (mehrere Historien)
# + Robust gegen HubSpot-Limits (auto split bei 400)
# + Seitenzahl unten (Footer): "Seite X von Y"

import os
import re
import textwrap
from io import BytesIO
from typing import Dict, List, Tuple

import pandas as pd
import requests
import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from shared.config import ConfigError, get_hubspot_token
from shared.hubspot import fetch_property_options, search_contacts_with_auto_split
from streamlit_ui import render_page_title


# ----------------------------
# Konfiguration
# ----------------------------
HISTORIE_PROPERTY = "historie"

HS_PROPERTIES = ["company", "firstname", "lastname", "expertenwissen", "vorqualifizierung"]

EXCEL_HEADERS = [
    "Unternehmensname",
    "Vorname",
    "Nachname",
    "Expertenwissen",
    "Herausforderung",
]
EXCEL_COLS = ["A", "B", "C", "D", "E"]

DEFAULT_WIDTHS = {"A": 28, "B": 18, "C": 18, "D": 26, "E": 30}

# Druck/Zeilenhöhe (Heuristik für Wrap-Text)
HEADER_ROW_HEIGHT = 22
BASE_ROW_HEIGHT = 15
LINE_HEIGHT = 15
ROW_PADDING = 2  # Sicherheitsmarge, damit Wrap-Text nicht abgeschnitten wird

# Start-Batchgröße (HubSpot kann bei zu vielen filterGroups 400 werfen -> auto-split)
INITIAL_FILTERGROUPS_PER_REQUEST = 5


# ----------------------------
# HubSpot Helpers (Requests)
# ----------------------------
def get_access_token() -> str:
    try:
        return get_hubspot_token(getattr(st, "secrets", None), os.environ)
    except ConfigError as exc:
        raise RuntimeError(
            "Kein HubSpot Token gefunden. Bitte HUBSPOT_TOKEN (oder HUBSPOT_ACCESS_TOKEN) setzen."
        ) from exc


@st.cache_data(ttl=3600)
def fetch_historie_options() -> List[Tuple[str, str]]:
    """
    Liefert [(label, value), ...] für das Property 'historie' aus der Property-Definition.
    Wenn Scopes fehlen oder Property Freitext ist: []
    """
    try:
        return fetch_property_options(
            HISTORIE_PROPERTY,
            token=get_access_token(),
            object_type="contacts",
        )
    except Exception:
        return []


def _chunks(lst: List[str], n: int) -> List[List[str]]:
    return [lst[i : i + n] for i in range(0, len(lst), n)]


def fetch_contacts_by_historien(historie_values: List[str]) -> List[dict]:
    """
    OR-Suche: Kontakte, deren 'historie' einen der Tokens enthält.
    HubSpot Search:
      - innerhalb einer filterGroup = AND
      - zwischen filterGroups = OR
    => pro Historie eine filterGroup mit einem Filter.
    Fix: Bei 400 automatisch splitten.
    """
    historie_values = [v.strip() for v in historie_values if v and v.strip()]
    historie_values = list(dict.fromkeys(historie_values))  # unique, Reihenfolge behalten
    if not historie_values:
        return []

    token = get_access_token()
    results_by_id: Dict[str, dict] = {}

    for batch in _chunks(historie_values, INITIAL_FILTERGROUPS_PER_REQUEST):
        filter_groups = [
            {
                "filters": [
                    {
                        "propertyName": HISTORIE_PROPERTY,
                        "operator": "CONTAINS_TOKEN",
                        "value": hv,
                    }
                ]
            }
            for hv in batch
        ]

        results = search_contacts_with_auto_split(
            filter_groups,
            HS_PROPERTIES,
            token=token,
        )

        for item in results:
            cid = item.get("id")
            if cid:
                results_by_id[cid] = item

    return list(results_by_id.values())


# ----------------------------
# Excel Helpers
# ----------------------------
def sanitize_filename(name: str) -> str:
    name = (name or "").strip()
    name = re.sub(r"[^\w\- ]+", "", name, flags=re.UNICODE)
    name = re.sub(r"\s+", "_", name)
    return name or "export"


def read_widths_from_template(uploaded_file) -> Dict[str, float]:
    """
    Liest Spaltenbreiten A–E aus einer hochgeladenen Excel (erstes Blatt).
    """
    try:
        wb = load_workbook(filename=BytesIO(uploaded_file.getvalue()))
        ws = wb.worksheets[0]
        widths: Dict[str, float] = {}
        for col in EXCEL_COLS:
            w = ws.column_dimensions[col].width
            if w:
                widths[col] = float(w)
        return widths
    except Exception:
        return {}


def estimate_wrapped_lines(text: str, col_width_excel: float) -> int:
    """
    Konservative Heuristik: Wie viele Zeilen braucht der Text ungefähr bei Wrap Text?

    Hinweis: OpenPyXL kann Excel-"AutoFit Row Height" (Font-Metriken) nicht exakt nachbilden.
    Deshalb rechnen wir bewusst etwas konservativ, damit beim Öffnen/Drucken nichts abgeschnitten wird.
    """
    if text is None:
        return 1

    s = str(text).replace("\r\n", "\n").replace("\r", "\n")
    if s.strip() == "":
        return 1

    # Excel-Spaltenbreite ist grob "Zeichen" in Standard-Font (Calibri 11).
    # Mit 0.95 werden wir etwas konservativer (-> mehr Zeilen).
    try:
        chars_per_line = max(6, int(float(col_width_excel) * 0.95))
    except Exception:
        chars_per_line = 12

    total_lines = 0
    for raw in s.split("\n"):
        line = raw.strip()
        if line == "":
            total_lines += 1
            continue

        wrapped = textwrap.wrap(
            line,
            width=chars_per_line,
            break_long_words=True,
            break_on_hyphens=True,
        )
        total_lines += max(1, len(wrapped))

    return max(1, total_lines)


def autofit_row_heights(ws, start_row: int, end_row: int, widths_final: Dict[str, float]) -> None:
    """
    Setzt Zeilenhöhen heuristisch basierend auf Inhalt + Spaltenbreite.
    """
    for r in range(start_row, end_row + 1):
        max_lines = 1
        for col in EXCEL_COLS:
            cell = ws[f"{col}{r}"]
            col_w = widths_final.get(col, DEFAULT_WIDTHS[col])
            max_lines = max(max_lines, estimate_wrapped_lines(cell.value, col_w))

        ws.row_dimensions[r].height = float(
            BASE_ROW_HEIGHT + (max_lines - 1) * LINE_HEIGHT + ROW_PADDING
        )


def build_excel_bytes(df: pd.DataFrame, list_title: str, widths: Dict[str, float]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Liste"

    header_font = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")
    thin = Side(style="thin")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header
    ws.append(EXCEL_HEADERS)
    ws.row_dimensions[1].height = HEADER_ROW_HEIGHT

    for c in range(1, len(EXCEL_HEADERS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.alignment = wrap
        cell.border = border_all

    # Daten
    for row in df.itertuples(index=False):
        ws.append(list(row))

    max_row = ws.max_row
    max_col = ws.max_column

    # Spaltenbreiten
    widths_final = DEFAULT_WIDTHS.copy()
    widths_final.update(widths or {})
    for col in EXCEL_COLS:
        ws.column_dimensions[col].width = widths_final.get(col, DEFAULT_WIDTHS[col])

    # Wrap + Rahmenlinien für alle Zellen
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = wrap
            cell.border = border_all

    # Auto-Zeilenhöhe (Heuristik)
    if max_row >= 2:
        autofit_row_heights(ws, start_row=2, end_row=max_row, widths_final=widths_final)

    # Druck-Setup
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.orientation = "portrait"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    # Wiederhole Header-Zeile (Spaltennamen) auf jeder Seite
    ws.print_title_rows = "1:1"

    # Titel oben (Header)
    title = (list_title or "").strip()
    if title:
        ws.oddHeader.center.text = f'&"Calibri,Bold"&14 {title}'
        ws.evenHeader.center.text = f'&"Calibri,Bold"&14 {title}'

    # Seitenzahl unten (Footer)
    ws.oddFooter.center.text = '&"Calibri"&10 Seite &P von &N'
    ws.evenFooter.center.text = '&"Calibri"&10 Seite &P von &N'

    # Print Area
    ws.print_area = f"A1:{get_column_letter(max_col)}{max_row}"

    # Ränder
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.6, bottom=0.6, header=0.3, footer=0.3)

    # Freeze Header
    ws.freeze_panes = "A2"

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


# ----------------------------
# Streamlit UI
# ----------------------------
st.set_page_config(page_title="HubSpot -> Druck-Excel", layout="wide")
render_page_title("HubSpot → Excel Export (druckfertig)")

options = fetch_historie_options()

historie_labels_selected: List[str] = []
historie_values_selected: List[str] = []

if not options:
    st.warning(
        "Ich konnte keine Dropdown-Optionen für 'historie' aus HubSpot laden.\n"
        "Entweder fehlen Scopes (crm.schemas.contacts.read) oder 'historie' ist ein Freitext-Feld.\n"
        "Gib mehrere Werte getrennt durch Komma/Zeilenumbruch ein."
    )
    raw = st.text_area(
        "Historie-Werte (mehrere möglich)",
        placeholder="z. B. 26DOR_TN, 26DOR_REF, ...",
    )
    historie_values_selected = [x.strip() for x in re.split(r"[,\n;]+", raw) if x.strip()]
    historie_labels_selected = historie_values_selected[:]
else:
    labels = [o[0] for o in options]
    label_to_value = {label: value for label, value in options}

    historie_labels_selected = st.multiselect(
        "Historie auswählen (mehrere möglich)",
        options=labels,
        default=[],
    )
    historie_values_selected = [label_to_value[lbl] for lbl in historie_labels_selected]

# Default Titel (aus den ersten 3 Historien)
if historie_labels_selected:
    default_title = ", ".join(historie_labels_selected[:3])
    if len(historie_labels_selected) > 3:
        default_title += f" (+{len(historie_labels_selected) - 3})"
else:
    default_title = ""

list_title_input = st.text_input(
    "Listentitel (optional, steht oben auf jedem gedruckten Blatt)",
    value=default_title,
)

template_file = st.file_uploader(
    "Excel-Template hochladen (optional) – Spaltenbreiten A–E werden übernommen",
    type=["xlsx"],
)

disabled_btn = len(historie_values_selected) == 0

if st.button("Excel erstellen", type="primary", disabled=disabled_btn):
    try:
        with st.spinner("Lade Daten aus HubSpot..."):
            contacts = fetch_contacts_by_historien(historie_values_selected)
    except requests.HTTPError as e:
        st.error("HubSpot API Fehler.\n\n" + str(e))
        st.stop()
    except Exception as e:
        st.error(f"Unerwarteter Fehler: {e}")
        st.stop()

    if not contacts:
        st.error("Keine Kontakte für diese Historie(n) gefunden.")
        st.stop()

    rows = []
    for c in contacts:
        p = c.get("properties", {}) or {}
        rows.append(
            {
                "Unternehmensname": (p.get("company") or "").strip(),
                "Vorname": (p.get("firstname") or "").strip(),
                "Nachname": (p.get("lastname") or "").strip(),
                "Expertenwissen": (p.get("expertenwissen") or "").strip(),
                "Herausforderung": (p.get("vorqualifizierung") or "").strip(),
            }
        )

    df = pd.DataFrame(rows, columns=EXCEL_HEADERS)

    # Sortierung nach Unternehmensname A-Z (case-insensitive)
    df["__sort_company"] = df["Unternehmensname"].fillna("").str.lower()
    df = df.sort_values("__sort_company", ascending=True).drop(columns=["__sort_company"])

    widths = {}
    if template_file is not None:
        widths = read_widths_from_template(template_file)

    excel_bytes = build_excel_bytes(df, list_title_input, widths)

    filename = f"{sanitize_filename(list_title_input or 'export')}.xlsx"
    st.success(f"Fertig. Kontakte: {len(df)}")
    st.download_button(
        label="Excel herunterladen",
        data=excel_bytes,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    st.dataframe(df, use_container_width=True)
