from __future__ import annotations

import re
import textwrap
from io import BytesIO

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

DEFAULT_EXCEL_HEADERS = [
    "Unternehmensname",
    "Vorname",
    "Nachname",
    "Expertenwissen",
    "Herausforderung",
]
EXCEL_HEADERS = DEFAULT_EXCEL_HEADERS
EXCEL_COLS = ["A", "B", "C", "D", "E"]

DEFAULT_WIDTHS = {"A": 28, "B": 18, "C": 18, "D": 26, "E": 30}

HEADER_ROW_HEIGHT = 22
BASE_ROW_HEIGHT = 15
LINE_HEIGHT = 15
ROW_PADDING = 2


def sanitize_filename(name: str) -> str:
    name = (name or "").strip()
    name = re.sub(r"[^\w\- ]+", "", name, flags=re.UNICODE)
    name = re.sub(r"\s+", "_", name)
    return name or "export"


def read_widths_from_excel_bytes(excel_bytes: bytes) -> dict[str, float]:
    """
    Read column widths A-E from the first worksheet of an Excel file.
    Returns an empty dict when the workbook cannot be read.
    """
    try:
        wb = load_workbook(filename=BytesIO(excel_bytes))
        ws = wb.worksheets[0]
        widths: dict[str, float] = {}
        for col in EXCEL_COLS:
            width = ws.column_dimensions[col].width
            if width:
                widths[col] = float(width)
        return widths
    except Exception:
        return {}


def estimate_wrapped_lines(text: str, col_width_excel: float) -> int:
    """
    Conservative heuristic for wrapped Excel text line count.
    """
    if text is None:
        return 1

    value = str(text).replace("\r\n", "\n").replace("\r", "\n")
    if value.strip() == "":
        return 1

    try:
        chars_per_line = max(6, int(float(col_width_excel) * 0.95))
    except Exception:
        chars_per_line = 12

    total_lines = 0
    for raw in value.split("\n"):
        line = raw.strip()
        if line == "":
            total_lines += 1
            continue

        wrapped = textwrap.wrap(
            line,
            width=chars_per_line,
            break_long_words=True,
            break_on_hyphens=True,
        )
        total_lines += max(1, len(wrapped))

    return max(1, total_lines)


def autofit_row_heights(ws, start_row: int, end_row: int, widths_final: dict[str, float]) -> None:
    """
    Set worksheet row heights heuristically based on content and column width.
    """
    for row_number in range(start_row, end_row + 1):
        max_lines = 1
        for col in EXCEL_COLS:
            cell = ws[f"{col}{row_number}"]
            col_width = widths_final.get(col, DEFAULT_WIDTHS[col])
            max_lines = max(max_lines, estimate_wrapped_lines(cell.value, col_width))

        ws.row_dimensions[row_number].height = float(
            BASE_ROW_HEIGHT + (max_lines - 1) * LINE_HEIGHT + ROW_PADDING
        )


def build_excel_bytes(
    df: pd.DataFrame,
    list_title: str,
    widths: dict[str, float],
    *,
    headers: list[str],
) -> bytes:
    if len(headers) != len(EXCEL_COLS):
        raise ValueError(f"Expected {len(EXCEL_COLS)} headers, got {len(headers)}.")

    export_headers = list(headers)
    export_df = df.reindex(columns=export_headers)

    wb = Workbook()
    ws = wb.active
    ws.title = "Liste"

    header_font = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")
    thin = Side(style="thin")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(export_headers)
    ws.row_dimensions[1].height = HEADER_ROW_HEIGHT

    for col_number in range(1, len(export_headers) + 1):
        cell = ws.cell(row=1, column=col_number)
        cell.font = header_font
        cell.alignment = wrap
        cell.border = border_all

    for row in export_df.itertuples(index=False):
        ws.append(list(row))

    max_row = ws.max_row
    max_col = ws.max_column

    widths_final = DEFAULT_WIDTHS.copy()
    widths_final.update(widths or {})
    for col in EXCEL_COLS:
        ws.column_dimensions[col].width = widths_final.get(col, DEFAULT_WIDTHS[col])

    for row_number in range(1, max_row + 1):
        for col_number in range(1, max_col + 1):
            cell = ws.cell(row=row_number, column=col_number)
            cell.alignment = wrap
            cell.border = border_all

    if max_row >= 2:
        autofit_row_heights(ws, start_row=2, end_row=max_row, widths_final=widths_final)

    ws.page_setup.paperSize = 9
    ws.page_setup.orientation = "portrait"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    ws.print_title_rows = "1:1"

    title = (list_title or "").strip()
    if title:
        ws.oddHeader.center.text = f'&"Calibri,Bold"&14 {title}'
        ws.evenHeader.center.text = f'&"Calibri,Bold"&14 {title}'

    ws.oddFooter.center.text = '&"Calibri"&10 Seite &P von &N'
    ws.evenFooter.center.text = '&"Calibri"&10 Seite &P von &N'

    ws.print_area = f"A1:{get_column_letter(max_col)}{max_row}"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.6, bottom=0.6, header=0.3, footer=0.3)
    ws.freeze_panes = "A2"

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
