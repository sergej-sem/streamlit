from __future__ import annotations

import io
from dataclasses import dataclass
from datetime import date, datetime

from openpyxl import load_workbook

from .parser import (
    SponsorRow,
    build_sponsor_row,
    cell_value,
    list_workbook_sheets,
    normalize_text,
    slugify,
)
from .planner import (
    DeadlineItem,
    build_deadlines,
)
from .rendering import build_html_body, build_subject, build_summary_dataframe


DEFAULT_SHEET_NAME = "Deals"
DEFAULT_EVENT_CITY = "Berlin"
DEFAULT_EVENT_START = date(2026, 5, 5)
DEFAULT_EVENT_END = date(2026, 5, 7)

NAME_COL = "B"


@dataclass(frozen=True)
class GeneratedMail:
    row_number: int
    sponsor_name: str
    language: str
    package: str
    to_email: str
    cc_email: str
    subject: str
    html_body: str
    html_file_name: str
    green_count: int
    red_count: int
    yellow_count: int
    white_count: int
    outlook_result: str = "not_requested"


@dataclass(frozen=True)
class GenerationResult:
    sheet_name: str
    event_city: str
    event_start: date
    event_end: date
    mails: tuple[GeneratedMail, ...]
    processed_count: int
    skipped_count: int


def parse_event_date(value: date | str) -> date:
    if isinstance(value, date):
        return value
    text = normalize_text(value)
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Ungültiges Datum: {value}")


def generate_deadline_mails(
    excel_bytes: bytes,
    sheet_name: str = DEFAULT_SHEET_NAME,
    event_city: str = DEFAULT_EVENT_CITY,
    event_start: date | str = DEFAULT_EVENT_START,
    event_end: date | str = DEFAULT_EVENT_END,
) -> GenerationResult:
    start_date = parse_event_date(event_start)
    end_date = parse_event_date(event_end)
    workbook = load_workbook(io.BytesIO(excel_bytes), data_only=False)

    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"Blatt '{sheet_name}' nicht gefunden. Verfügbar: {', '.join(workbook.sheetnames)}"
            )

        ws = workbook[sheet_name]
        mails: list[GeneratedMail] = []
        candidate_rows = 0

        for row_number in range(2, ws.max_row + 1):
            sponsor_name = normalize_text(cell_value(ws, NAME_COL, row_number))
            if not sponsor_name:
                continue

            candidate_rows += 1
            sponsor = build_sponsor_row(ws, row_number)
            if sponsor is None:
                continue

            items = build_deadlines(ws, sponsor)
            subject = build_subject(sponsor.language, sponsor.sponsor_name)
            html_body = build_html_body(sponsor, items, event_city, start_date, end_date)
            html_file_name = f"{row_number:03d}_{slugify(sponsor.sponsor_name)}.html"
            green_count = sum(1 for item in items if item.status == "green")
            red_count = sum(1 for item in items if item.status == "red")
            yellow_count = sum(1 for item in items if item.status == "yellow")
            white_count = sum(1 for item in items if item.status == "white")

            mails.append(
                GeneratedMail(
                    row_number=sponsor.row_number,
                    sponsor_name=sponsor.sponsor_name,
                    language=sponsor.language,
                    package=sponsor.package,
                    to_email=sponsor.to_email,
                    cc_email=sponsor.cc_email,
                    subject=subject,
                    html_body=html_body,
                    html_file_name=html_file_name,
                    green_count=green_count,
                    red_count=red_count,
                    yellow_count=yellow_count,
                    white_count=white_count,
                )
            )

        return GenerationResult(
            sheet_name=sheet_name,
            event_city=event_city,
            event_start=start_date,
            event_end=end_date,
            mails=tuple(mails),
            processed_count=len(mails),
            skipped_count=max(candidate_rows - len(mails), 0),
        )
    finally:
        workbook.close()
