from __future__ import annotations

import io
import re
from dataclasses import dataclass

from openpyxl import load_workbook


NAME_COL = "B"
PACKAGE_COL = "E"
LANG_COL = "K"
CONTACT1_FIRST_COL = "L"
CONTACT1_LAST_COL = "M"
CONTACT1_EMAIL_COL = "N"
CONTACT2_FIRST_COL = "O"
CONTACT2_LAST_COL = "P"
CONTACT2_EMAIL_COL = "Q"
DEAL_ACTIVE_COL = "D"

ENGLISH_LANG_VALUES = {"ENG", "EN", "ENGLISH"}
ACTIVE_MARKERS = {"check", "x", "ja", "yes", "true", "ok", "done", "erhalten"}


@dataclass(frozen=True)
class SponsorRow:
    row_number: int
    sponsor_name: str
    package: str
    language: str
    to_email: str
    cc_email: str
    contact_first_name: str
    contact_last_name: str


def list_workbook_sheets(excel_bytes: bytes) -> list[str]:
    workbook = load_workbook(io.BytesIO(excel_bytes), read_only=True, data_only=False)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def cell_value(ws, col: str, row: int) -> object:
    return ws[f"{col}{row}"].value


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_lang(value: object) -> str:
    return "EN" if normalize_text(value).upper() in ENGLISH_LANG_VALUES else "DE"


def normalize_package(value: object) -> str:
    return normalize_text(value)


def slugify(value: str) -> str:
    text = normalize_text(value).casefold()
    text = re.sub(r"[^\w]+", "_", text, flags=re.UNICODE)
    text = text.strip("_")
    return text or "sponsor"


def build_sponsor_row(ws, row: int) -> SponsorRow | None:
    sponsor_name = normalize_text(cell_value(ws, NAME_COL, row))
    if not sponsor_name:
        return None

    deal_active = normalize_text(cell_value(ws, DEAL_ACTIVE_COL, row)).lower()
    if deal_active and deal_active not in ACTIVE_MARKERS:
        return None

    package = normalize_package(cell_value(ws, PACKAGE_COL, row))
    language = normalize_lang(cell_value(ws, LANG_COL, row))

    first_name = normalize_text(cell_value(ws, CONTACT1_FIRST_COL, row))
    last_name = normalize_text(cell_value(ws, CONTACT1_LAST_COL, row))
    to_email = normalize_text(cell_value(ws, CONTACT1_EMAIL_COL, row))
    cc_email = normalize_text(cell_value(ws, CONTACT2_EMAIL_COL, row))

    if not to_email and cc_email:
        to_email = cc_email
        cc_email = ""
        first_name = normalize_text(cell_value(ws, CONTACT2_FIRST_COL, row))
        last_name = normalize_text(cell_value(ws, CONTACT2_LAST_COL, row))

    if not to_email:
        return None

    return SponsorRow(
        row_number=row,
        sponsor_name=sponsor_name,
        package=package,
        language=language,
        to_email=to_email,
        cc_email=cc_email,
        contact_first_name=first_name,
        contact_last_name=last_name,
    )
