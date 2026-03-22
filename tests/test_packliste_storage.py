import os
import pickle
import tempfile
import time
import unittest
from pathlib import Path

import openpyxl
import pandas as pd

from packliste.storage import (
    _read_sidecar,
    _sidecar_newer,
    _sidecar_path,
    _write_sidecar,
    load_df,
    load_last_active_path,
    restore_state_for_path,
    save_df,
    save_last_active_path,
)


SHEET_NAME = "Packliste"
REQUIRED_COLS = ("Bereich", "Gegenstand", "Beschreibung", "Verpackt")
CHECKBOX_COLS = ("Verpackt", "Verladen")
CHECKED = "Y"
UNCHECKED = "N"


def _make_workbook(path: Path, rows: list[dict], *, include_verladen: bool = False) -> None:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = SHEET_NAME
    worksheet["A1"] = "Packliste Test"
    worksheet["A2"] = "Generated"

    headers = list(REQUIRED_COLS)
    if include_verladen:
        headers.append("Verladen")

    for col_idx, header in enumerate(headers, start=1):
        worksheet.cell(row=3, column=col_idx, value=header)

    for row_idx, row in enumerate(rows, start=4):
        for col_idx, header in enumerate(headers, start=1):
            value = row.get(header, "")
            if header in CHECKBOX_COLS:
                value = CHECKED if bool(value) else UNCHECKED
            worksheet.cell(row=row_idx, column=col_idx, value=value if value != "" else None)

    workbook.save(path)
    workbook.close()


class PacklisteStorageTests(unittest.TestCase):
    def setUp(self) -> None:
        self.tmpdir = tempfile.TemporaryDirectory()
        self.addCleanup(self.tmpdir.cleanup)
        self.base = Path(self.tmpdir.name)

    def test_meta_roundtrip_returns_existing_path(self) -> None:
        excel_path = self.base / "packliste.xlsx"
        _make_workbook(excel_path, [])
        meta_file = self.base / "meta.json"

        save_last_active_path(meta_file, str(excel_path))

        self.assertEqual(load_last_active_path(meta_file), str(excel_path))

    def test_meta_returns_empty_for_missing_target(self) -> None:
        meta_file = self.base / "meta.json"

        save_last_active_path(meta_file, str(self.base / "missing.xlsx"))

        self.assertEqual(load_last_active_path(meta_file), "")

    def test_meta_returns_empty_for_directory_target(self) -> None:
        meta_file = self.base / "meta.json"

        save_last_active_path(meta_file, str(self.base))

        self.assertEqual(load_last_active_path(meta_file), "")

    def test_meta_returns_empty_for_non_xlsx_file(self) -> None:
        meta_file = self.base / "meta.json"
        text_file = self.base / "packliste.txt"
        text_file.write_text("not excel", encoding="utf-8")

        save_last_active_path(meta_file, str(text_file))

        self.assertEqual(load_last_active_path(meta_file), "")

    def test_sidecar_roundtrip_reads_dataframe(self) -> None:
        excel_path = str(self.base / "packliste.xlsx")
        df = pd.DataFrame([{"Bereich": "Bar", "Verpackt": True}])

        _write_sidecar(df, excel_path)

        loaded = _read_sidecar(excel_path)
        pd.testing.assert_frame_equal(loaded, df)

    def test_sidecar_path_mismatch_is_ignored(self) -> None:
        excel_path = str(self.base / "packliste.xlsx")
        sidecar = _sidecar_path(excel_path)
        payload = {
            "df": pd.DataFrame([{"Bereich": "Bar"}]),
            "path": str(self.base / "other.xlsx"),
            "ts": time.time(),
        }

        with open(sidecar, "wb") as fh:
            pickle.dump(payload, fh, protocol=pickle.HIGHEST_PROTOCOL)

        self.assertIsNone(_read_sidecar(excel_path))

    def test_restore_prefers_newer_sidecar(self) -> None:
        excel_path = self.base / "packliste.xlsx"
        _make_workbook(
            excel_path,
            [
                {
                    "Bereich": "Bar",
                    "Gegenstand": "Glas",
                    "Beschreibung": "Klein",
                    "Verpackt": False,
                }
            ],
        )
        sidecar_df = pd.DataFrame(
            [
                {
                    "Bereich": "Bar",
                    "Gegenstand": "Glas",
                    "Beschreibung": "Aktuell",
                    "Verpackt": True,
                    "Verladen": False,
                }
            ]
        )
        _write_sidecar(sidecar_df, str(excel_path))

        now = time.time()
        os.utime(excel_path, (now - 10, now - 10))
        os.utime(_sidecar_path(str(excel_path)), (now, now))

        restored_df, source = restore_state_for_path(
            str(excel_path),
            sheet_name=SHEET_NAME,
            required_cols=REQUIRED_COLS,
            checkbox_cols=CHECKBOX_COLS,
            checked_value=CHECKED,
        )

        self.assertEqual(source, "sidecar")
        pd.testing.assert_frame_equal(restored_df, sidecar_df)
        self.assertTrue(_sidecar_newer(str(excel_path)))

    def test_restore_falls_back_to_excel_when_sidecar_not_newer(self) -> None:
        excel_path = self.base / "packliste.xlsx"
        _make_workbook(
            excel_path,
            [
                {
                    "Bereich": "Lager",
                    "Gegenstand": "Box",
                    "Beschreibung": "Gross",
                    "Verpackt": True,
                }
            ],
        )
        stale_df = pd.DataFrame(
            [
                {
                    "Bereich": "Alt",
                    "Gegenstand": "Alt",
                    "Beschreibung": "Alt",
                    "Verpackt": False,
                }
            ]
        )
        _write_sidecar(stale_df, str(excel_path))

        now = time.time()
        os.utime(_sidecar_path(str(excel_path)), (now - 20, now - 20))
        os.utime(excel_path, (now, now))

        restored_df, source = restore_state_for_path(
            str(excel_path),
            sheet_name=SHEET_NAME,
            required_cols=REQUIRED_COLS,
            checkbox_cols=CHECKBOX_COLS,
            checked_value=CHECKED,
        )

        self.assertEqual(source, "excel")
        self.assertEqual(restored_df.loc[0, "Bereich"], "Lager")
        self.assertTrue(bool(restored_df.loc[0, "Verpackt"]))
        self.assertIn("Verladen", restored_df.columns)
        self.assertFalse(bool(restored_df.loc[0, "Verladen"]))

    def test_load_df_reads_header_row_and_checkbox_columns(self) -> None:
        excel_path = self.base / "packliste.xlsx"
        _make_workbook(
            excel_path,
            [
                {
                    "Bereich": "Bar",
                    "Gegenstand": "Glas",
                    "Beschreibung": "6x",
                    "Verpackt": True,
                }
            ],
        )

        df = load_df(
            str(excel_path),
            sheet_name=SHEET_NAME,
            required_cols=REQUIRED_COLS,
            checkbox_cols=CHECKBOX_COLS,
            checked_value=CHECKED,
        )

        self.assertEqual(list(df.columns), list(REQUIRED_COLS) + ["Verladen"])
        self.assertTrue(bool(df.loc[0, "Verpackt"]))
        self.assertFalse(bool(df.loc[0, "Verladen"]))

    def test_save_df_roundtrip_updates_excel_and_adds_verladen_column(self) -> None:
        excel_path = self.base / "packliste.xlsx"
        _make_workbook(
            excel_path,
            [
                {
                    "Bereich": "Bar",
                    "Gegenstand": "Glas",
                    "Beschreibung": "6x",
                    "Verpackt": False,
                }
            ],
            include_verladen=False,
        )

        df = load_df(
            str(excel_path),
            sheet_name=SHEET_NAME,
            required_cols=REQUIRED_COLS,
            checkbox_cols=CHECKBOX_COLS,
            checked_value=CHECKED,
        )
        df.loc[0, "Beschreibung"] = "12x"
        df.loc[0, "Verpackt"] = True
        df.loc[0, "Verladen"] = True

        save_df(
            df,
            str(excel_path),
            sheet_name=SHEET_NAME,
            required_cols=REQUIRED_COLS,
            checkbox_cols=CHECKBOX_COLS,
            checked_value=CHECKED,
            unchecked_value=UNCHECKED,
        )

        workbook = openpyxl.load_workbook(excel_path)
        worksheet = workbook[SHEET_NAME]
        headers = [worksheet.cell(row=3, column=col_idx).value for col_idx in range(1, 6)]
        row_values = [worksheet.cell(row=4, column=col_idx).value for col_idx in range(1, 6)]
        workbook.close()

        self.assertEqual(headers, ["Bereich", "Gegenstand", "Beschreibung", "Verpackt", "Verladen"])
        self.assertEqual(row_values, ["Bar", "Glas", "12x", CHECKED, CHECKED])


if __name__ == "__main__":
    unittest.main()
