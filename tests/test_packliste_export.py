from io import BytesIO
import unittest

import openpyxl
import pandas as pd

from packliste.export import download_filename, to_excel_bytes


class PacklisteExportTests(unittest.TestCase):
    def test_to_excel_bytes_returns_openable_workbook(self) -> None:
        df = pd.DataFrame(
            [
                {
                    "Bereich": "Bar",
                    "Beschreibung": "6x Glaser",
                    "Verpackt": True,
                    "Verladen": False,
                    "Fortschritt": 75,
                }
            ]
        )

        excel_bytes = to_excel_bytes(
            df,
            sheet_name="Packliste",
            checkbox_cols=("Verpackt", "Verladen"),
        )

        self.assertTrue(excel_bytes)

        workbook = openpyxl.load_workbook(BytesIO(excel_bytes))
        worksheet = workbook["Packliste"]

        headers = [worksheet.cell(row=1, column=idx).value for idx in range(1, 6)]
        values = [worksheet.cell(row=2, column=idx).value for idx in range(1, 6)]
        progress_cell = worksheet["E2"]

        workbook.close()

        self.assertEqual(
            headers,
            ["Bereich", "Beschreibung", "Verpackt", "Verladen", "Fortschritt"],
        )
        self.assertEqual(values, ["Bar", "6x Glaser", "Ja", "Nein", 75])
        self.assertEqual(progress_cell.number_format, '0" %"')

    def test_download_filename_uses_default_name_when_path_missing(self) -> None:
        self.assertEqual(download_filename(None, "gesamt"), "packliste_gesamt.xlsx")

    def test_download_filename_uses_stem_from_path(self) -> None:
        self.assertEqual(
            download_filename(r"C:\tmp\liste.xlsx", "gesamt"),
            "liste_gesamt.xlsx",
        )


if __name__ == "__main__":
    unittest.main()
