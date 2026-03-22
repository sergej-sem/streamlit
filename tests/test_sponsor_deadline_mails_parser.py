from __future__ import annotations

import io
import unittest

from openpyxl import Workbook

from sponsor_deadline_mails.parser import (
    build_sponsor_row,
    list_workbook_sheets,
    normalize_lang,
    normalize_package,
    slugify,
)


def _workbook_bytes_with_sheets(*sheet_names: str) -> bytes:
    workbook = Workbook()
    workbook.active.title = sheet_names[0]
    for name in sheet_names[1:]:
        workbook.create_sheet(name)
    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def _make_sponsor_sheet():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Deals"
    return workbook, worksheet


class SponsorDeadlineMailParserTests(unittest.TestCase):
    def test_list_workbook_sheets_reads_sheet_names(self) -> None:
        excel_bytes = _workbook_bytes_with_sheets("Deals", "Archive", "Extra")

        self.assertEqual(list_workbook_sheets(excel_bytes), ["Deals", "Archive", "Extra"])

    def test_normalize_lang_maps_english_values(self) -> None:
        self.assertEqual(normalize_lang("ENG"), "EN")
        self.assertEqual(normalize_lang("ENGLISH"), "EN")
        self.assertEqual(normalize_lang("en"), "EN")
        self.assertEqual(normalize_lang("DE"), "DE")
        self.assertEqual(normalize_lang(""), "DE")

    def test_normalize_package_only_trims(self) -> None:
        self.assertEqual(normalize_package("  Gold  "), "Gold")

    def test_slugify_normalizes_and_falls_back(self) -> None:
        self.assertEqual(slugify("Alpha & Beta"), "alpha_beta")
        self.assertEqual(slugify("   "), "sponsor")

    def test_build_sponsor_row_returns_active_valid_sponsor(self) -> None:
        workbook, worksheet = _make_sponsor_sheet()
        self.addCleanup(workbook.close)
        worksheet["B2"] = "Acme GmbH"
        worksheet["D2"] = "ja"
        worksheet["E2"] = "Gold"
        worksheet["K2"] = "ENG"
        worksheet["L2"] = "Alice"
        worksheet["M2"] = "Smith"
        worksheet["N2"] = "alice@example.com"
        worksheet["Q2"] = "copy@example.com"

        sponsor = build_sponsor_row(worksheet, 2)

        self.assertIsNotNone(sponsor)
        self.assertEqual(sponsor.sponsor_name, "Acme GmbH")
        self.assertEqual(sponsor.language, "EN")
        self.assertEqual(sponsor.package, "Gold")
        self.assertEqual(sponsor.to_email, "alice@example.com")
        self.assertEqual(sponsor.cc_email, "copy@example.com")
        self.assertEqual(sponsor.contact_first_name, "Alice")

    def test_build_sponsor_row_falls_back_to_second_contact(self) -> None:
        workbook, worksheet = _make_sponsor_sheet()
        self.addCleanup(workbook.close)
        worksheet["B2"] = "Fallback AG"
        worksheet["D2"] = "yes"
        worksheet["E2"] = "Platinum"
        worksheet["K2"] = "DE"
        worksheet["O2"] = "Bernd"
        worksheet["P2"] = "Beispiel"
        worksheet["Q2"] = "bernd@example.com"

        sponsor = build_sponsor_row(worksheet, 2)

        self.assertIsNotNone(sponsor)
        self.assertEqual(sponsor.to_email, "bernd@example.com")
        self.assertEqual(sponsor.cc_email, "")
        self.assertEqual(sponsor.contact_first_name, "Bernd")
        self.assertEqual(sponsor.contact_last_name, "Beispiel")

    def test_build_sponsor_row_ignores_sponsor_without_mail(self) -> None:
        workbook, worksheet = _make_sponsor_sheet()
        self.addCleanup(workbook.close)
        worksheet["B2"] = "No Mail GmbH"
        worksheet["D2"] = "ja"

        self.assertIsNone(build_sponsor_row(worksheet, 2))

    def test_build_sponsor_row_ignores_inactive_deal(self) -> None:
        workbook, worksheet = _make_sponsor_sheet()
        self.addCleanup(workbook.close)
        worksheet["B2"] = "Inactive GmbH"
        worksheet["D2"] = "no"
        worksheet["N2"] = "inactive@example.com"

        self.assertIsNone(build_sponsor_row(worksheet, 2))


if __name__ == "__main__":
    unittest.main()
