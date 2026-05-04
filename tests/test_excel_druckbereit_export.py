from __future__ import annotations

import unittest
from io import BytesIO

import pandas as pd
from openpyxl import Workbook, load_workbook

from excel_druckbereit_generator.export import (
    BASE_ROW_HEIGHT,
    EXCEL_HEADERS,
    build_excel_bytes,
    read_widths_from_excel_bytes,
    sanitize_filename,
)


class SanitizeFilenameTests(unittest.TestCase):
    def test_sanitize_filename_normalizes_whitespace_and_symbols(self) -> None:
        self.assertEqual(sanitize_filename("  Hallo Welt!  "), "Hallo_Welt")

    def test_sanitize_filename_falls_back_to_export(self) -> None:
        self.assertEqual(sanitize_filename("!!!"), "export")


class ReadWidthsTests(unittest.TestCase):
    def test_read_widths_from_excel_bytes_reads_first_sheet_widths(self) -> None:
        workbook = Workbook()
        ws = workbook.active
        ws.column_dimensions["A"].width = 11
        ws.column_dimensions["C"].width = 23.5

        bio = BytesIO()
        workbook.save(bio)

        widths = read_widths_from_excel_bytes(bio.getvalue())

        self.assertEqual(widths["A"], 11.0)
        self.assertEqual(widths["C"], 23.5)

    def test_read_widths_from_excel_bytes_returns_empty_dict_for_invalid_bytes(self) -> None:
        self.assertEqual(read_widths_from_excel_bytes(b"not-an-excel-file"), {})


class BuildExcelBytesTests(unittest.TestCase):
    def test_build_excel_bytes_creates_expected_workbook(self) -> None:
        df = pd.DataFrame(
            [
                {
                    "Unternehmensname": "Acme GmbH",
                    "Vorname": "Ada",
                    "Nachname": "Lovelace",
                    "Expertenwissen": "Cloud",
                    "Herausforderung": "Migration",
                }
            ],
            columns=EXCEL_HEADERS,
        )

        data = build_excel_bytes(df, "Eventliste", {"A": 42, "E": 16}, headers=EXCEL_HEADERS)

        self.assertTrue(data)

        workbook = load_workbook(BytesIO(data))
        ws = workbook["Liste"]

        self.assertEqual([ws.cell(1, i).value for i in range(1, 6)], EXCEL_HEADERS)
        self.assertEqual(ws["A2"].value, "Acme GmbH")
        self.assertEqual(ws["B2"].value, "Ada")
        self.assertEqual(ws["E2"].value, "Migration")
        self.assertEqual(ws.freeze_panes, "A2")
        self.assertEqual(ws.column_dimensions["A"].width, 42.0)
        self.assertEqual(ws.column_dimensions["E"].width, 16.0)
        self.assertIn("Eventliste", ws.oddHeader.center.text)
        self.assertIn("Seite &P von &N", ws.oddFooter.center.text)

    def test_build_excel_bytes_sets_larger_row_height_for_wrapped_content(self) -> None:
        df = pd.DataFrame(
            [
                {
                    "Unternehmensname": "Acme GmbH",
                    "Vorname": "Ada",
                    "Nachname": "Lovelace",
                    "Expertenwissen": "Cloud",
                    "Herausforderung": "Sehr langer Text " * 10,
                }
            ],
            columns=EXCEL_HEADERS,
        )

        data = build_excel_bytes(df, "", {"E": 10}, headers=EXCEL_HEADERS)
        workbook = load_workbook(BytesIO(data))
        ws = workbook["Liste"]

        self.assertGreater(float(ws.row_dimensions[2].height), float(BASE_ROW_HEIGHT))

    def test_build_excel_bytes_uses_custom_headers_in_requested_order(self) -> None:
        custom_headers = [
            "Unternehmensname",
            "Jobtitel",
            "Vorname",
            "Nachname",
            "Datensatz ID",
        ]
        df = pd.DataFrame(
            [
                {
                    "Datensatz ID": "4711",
                    "Nachname": "Lovelace",
                    "Vorname": "Ada",
                    "Jobtitel": "CTO",
                    "Unternehmensname": "Acme GmbH",
                }
            ],
            columns=["Datensatz ID", "Nachname", "Vorname", "Jobtitel", "Unternehmensname"],
        )

        data = build_excel_bytes(df, "Eventliste", {"B": 24}, headers=custom_headers)
        workbook = load_workbook(BytesIO(data))
        ws = workbook["Liste"]

        self.assertEqual([ws.cell(1, i).value for i in range(1, 6)], custom_headers)
        self.assertEqual([ws.cell(2, i).value for i in range(1, 6)], ["Acme GmbH", "CTO", "Ada", "Lovelace", "4711"])


if __name__ == "__main__":
    unittest.main()
