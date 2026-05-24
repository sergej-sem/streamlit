from __future__ import annotations

import io
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from serienmailing.eventende import (
    assemble_eventende_sponsors,
    build_password_protected_excel_attachment,
    build_eventende_serienmails,
    build_eventende_summary_dataframe,
    inspect_eventende_sources,
)


def _make_workbook_bytes(rows: list[dict]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Deals"
    for index, row in enumerate(rows, start=2):
        worksheet[f"B{index}"] = row.get("sponsor_name", "")
        worksheet[f"D{index}"] = row.get("active", "")
        worksheet[f"E{index}"] = row.get("package", "")
        worksheet[f"K{index}"] = row.get("language", "")
        worksheet[f"L{index}"] = row.get("first_name", "")
        worksheet[f"M{index}"] = row.get("last_name", "")
        worksheet[f"N{index}"] = row.get("to_email", "")
        worksheet[f"O{index}"] = row.get("cc_first_name", "")
        worksheet[f"P{index}"] = row.get("cc_last_name", "")
        worksheet[f"Q{index}"] = row.get("cc_email", "")
    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


class _EventEndFixture:
    def __init__(self, workbook_rows: list[dict]) -> None:
        self._tmp = tempfile.TemporaryDirectory()
        self.root = Path(self._tmp.name)
        self.workbook_path = self.root / "00_Master_Sponsoren_Infos.xlsx"
        self.kontaktliste_path = self.root / "Kontaktliste.xlsx"
        self.gespraechsplaene_dir = self.root / "Gesprächspläne"
        self.vortragslisten_dir = self.root / "Vortragslisten"
        self.gespraechsplaene_dir.mkdir()
        self.vortragslisten_dir.mkdir()
        self.workbook_path.write_bytes(_make_workbook_bytes(workbook_rows))

    def cleanup(self) -> None:
        self._tmp.cleanup()

    def write_kontaktliste(self) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Teilnehmerliste"
        worksheet["A1"] = "Company"
        worksheet["B1"] = "Vorname"
        worksheet["A2"] = "Acme"
        worksheet["B2"] = "Alice"
        workbook.save(self.kontaktliste_path)
        workbook.close()

    def write_gespraechsplan(self, filename: str) -> None:
        (self.gespraechsplaene_dir / filename).write_bytes(filename.encode("utf-8"))

    def write_vortragsliste(self, filename: str) -> None:
        (self.vortragslisten_dir / filename).write_bytes(filename.encode("utf-8"))

    @staticmethod
    def fake_kontaktliste_builder(path: Path, password: str):
        return build_password_protected_excel_attachment(
            path,
            password,
            protector=lambda src, dst, _password: dst.write_bytes(src.read_bytes() + b"::protected"),
        )

    def assemble(self, **kwargs):
        kwargs.setdefault("kontaktliste_attachment_builder", self.fake_kontaktliste_builder)
        return assemble_eventende_sponsors(
            workbook_path=self.workbook_path,
            kontaktliste_path=self.kontaktliste_path,
            gespraechsplaene_dir=self.gespraechsplaene_dir,
            vortragslisten_dir=self.vortragslisten_dir,
            **kwargs,
        )


class EventEndSourceStatusTests(unittest.TestCase):
    def test_inspect_eventende_sources_reports_counts_and_missing_paths(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            workbook_path = root / "00_Master_Sponsoren_Infos.xlsx"
            kontaktliste_path = root / "Kontaktliste.xlsx"
            gespraechsplaene_dir = root / "Gesprächspläne"
            vortragslisten_dir = root / "Vortragslisten"
            workbook_path.write_bytes(b"workbook")
            gespraechsplaene_dir.mkdir()
            (gespraechsplaene_dir / "A.pdf").write_bytes(b"%PDF")

            status = inspect_eventende_sources(
                workbook_path=workbook_path,
                kontaktliste_path=kontaktliste_path,
                gespraechsplaene_dir=gespraechsplaene_dir,
                vortragslisten_dir=vortragslisten_dir,
            )

            self.assertTrue(status.workbook_exists)
            self.assertFalse(status.kontaktliste_exists)
            self.assertTrue(status.gespraechsplaene_exists)
            self.assertEqual(1, status.gespraechsplaene_count)
            self.assertFalse(status.vortragslisten_exists)
            self.assertEqual(0, status.vortragslisten_count)


class ContactListProtectionTests(unittest.TestCase):
    def test_build_password_protected_excel_attachment_returns_attachment(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            source_path = Path(tmpdir) / "Kontaktliste.xlsx"
            workbook = Workbook()
            workbook.active["A1"] = "test"
            workbook.save(source_path)
            workbook.close()

            calls = []

            def fake_protector(src: Path, dst: Path, password: str) -> None:
                calls.append((src, dst, password))
                dst.write_bytes(src.read_bytes() + b"::protected")

            attachment = build_password_protected_excel_attachment(
                source_path,
                "Secret123",
                protector=fake_protector,
            )

            self.assertEqual("Kontaktliste.xlsx", attachment.filename)
            self.assertTrue(attachment.content.endswith(b"::protected"))
            self.assertEqual(1, len(calls))
            self.assertEqual(source_path, calls[0][0])
            self.assertEqual("Secret123", calls[0][2])

    def test_build_password_protected_excel_attachment_requires_output(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            source_path = Path(tmpdir) / "Kontaktliste.xlsx"
            workbook = Workbook()
            workbook.active["A1"] = "test"
            workbook.save(source_path)
            workbook.close()

            with self.assertRaisesRegex(RuntimeError, "nicht erzeugt"):
                build_password_protected_excel_attachment(
                    source_path,
                    "Secret123",
                    protector=lambda src, dst, password: None,
                )


class AssembleEventEndSponsorsTests(unittest.TestCase):
    def test_includes_only_supported_active_packages(self) -> None:
        fixture = _EventEndFixture(
            [
                {
                    "sponsor_name": "Alpha Premium",
                    "active": "Check",
                    "package": "Premium",
                    "first_name": "Alice",
                    "to_email": "alice@example.com",
                },
                {
                    "sponsor_name": "Beta Gold",
                    "active": "Check",
                    "package": "Gold",
                    "first_name": "Bob",
                    "to_email": "bob@example.com",
                },
                {
                    "sponsor_name": "Gamma Bronze",
                    "active": "Check",
                    "package": "Bronze",
                    "first_name": "Cara",
                    "to_email": "cara@example.com",
                },
                {
                    "sponsor_name": "Delta Platin",
                    "active": "no",
                    "package": "Platin",
                    "first_name": "Dora",
                    "to_email": "dora@example.com",
                },
            ]
        )
        self.addCleanup(fixture.cleanup)
        fixture.write_kontaktliste()
        fixture.write_gespraechsplan("Alpha Premium_Gesprächsplan.pdf")
        fixture.write_gespraechsplan("Beta Gold_Gesprächsplan.pdf")
        fixture.write_vortragsliste("Beta Gold_Vortragsliste.pdf")

        result = fixture.assemble()

        self.assertEqual(["Alpha Premium", "Beta Gold"], [sponsor.sponsor_name for sponsor in result.sponsors])
        self.assertEqual(2, result.ready_count)
        self.assertEqual(0, result.blocked_count)

    def test_falls_back_to_second_contact_when_first_email_missing(self) -> None:
        fixture = _EventEndFixture(
            [
                {
                    "sponsor_name": "Fallback Gold",
                    "active": "Check",
                    "package": "Gold",
                    "cc_first_name": "Bernd",
                    "cc_last_name": "Beispiel",
                    "cc_email": "bernd@example.com",
                }
            ]
        )
        self.addCleanup(fixture.cleanup)
        fixture.write_kontaktliste()
        fixture.write_gespraechsplan("Fallback Gold_Gesprächsplan.pdf")
        fixture.write_vortragsliste("Fallback Gold_Vortragsliste.pdf")

        result = fixture.assemble()

        sponsor = result.sponsors[0]
        self.assertEqual("bernd@example.com", sponsor.to_email)
        self.assertEqual("", sponsor.cc_email)
        self.assertTrue(sponsor.is_ready)

    def test_premium_requires_pdf_gespraechsplan(self) -> None:
        fixture = _EventEndFixture(
            [
                {
                    "sponsor_name": "Müller & Söhne GmbH",
                    "active": "Check",
                    "package": "Premium",
                    "first_name": "Mila",
                    "to_email": "mila@example.com",
                }
            ]
        )
        self.addCleanup(fixture.cleanup)
        fixture.write_kontaktliste()
        fixture.write_gespraechsplan("Müller & Söhne GmbH_Gesprächsplan.pdf")

        result = fixture.assemble()

        sponsor = result.sponsors[0]
        self.assertTrue(sponsor.is_ready)
        self.assertEqual(
            ("Kontaktliste.xlsx", "Müller & Söhne GmbH_Gesprächsplan.pdf"),
            sponsor.attachment_names,
        )

    def test_gold_requires_pdf_gespraechsplan_and_pdf_vortragsliste(self) -> None:
        fixture = _EventEndFixture(
            [
                {
                    "sponsor_name": "Acme Gold",
                    "active": "Check",
                    "package": "Gold",
                    "first_name": "Gina",
                    "to_email": "gina@example.com",
                }
            ]
        )
        self.addCleanup(fixture.cleanup)
        fixture.write_kontaktliste()
        fixture.write_gespraechsplan("Acme Gold_Gesprächsplan.pdf")
        fixture.write_vortragsliste("Acme Gold_Vortragsliste.pdf")

        result = fixture.assemble()

        sponsor = result.sponsors[0]
        self.assertTrue(sponsor.is_ready)
        self.assertEqual(
            ("Kontaktliste.xlsx", "Acme Gold_Gesprächsplan.pdf", "Acme Gold_Vortragsliste.pdf"),
            sponsor.attachment_names,
        )

    def test_missing_required_file_blocks_only_that_sponsor(self) -> None:
        fixture = _EventEndFixture(
            [
                {
                    "sponsor_name": "Ready Premium",
                    "active": "Check",
                    "package": "Premium",
                    "first_name": "Ria",
                    "to_email": "ria@example.com",
                },
                {
                    "sponsor_name": "Blocked Gold",
                    "active": "Check",
                    "package": "Gold",
                    "first_name": "Ben",
                    "to_email": "ben@example.com",
                },
            ]
        )
        self.addCleanup(fixture.cleanup)
        fixture.write_kontaktliste()
        fixture.write_gespraechsplan("Ready Premium_Gesprächsplan.pdf")

        result = fixture.assemble()

        self.assertEqual(1, result.ready_count)
        self.assertEqual(1, result.blocked_count)
        self.assertTrue(result.sponsors[0].is_ready)
        self.assertFalse(result.sponsors[1].is_ready)
        self.assertIn("Gesprächsplan fehlt", result.sponsors[1].details)
        self.assertIn("Vortragsliste fehlt", result.sponsors[1].details)

    def test_contactlist_protection_failure_blocks_all_sponsors(self) -> None:
        fixture = _EventEndFixture(
            [
                {
                    "sponsor_name": "Ready Premium",
                    "active": "Check",
                    "package": "Premium",
                    "first_name": "Ria",
                    "to_email": "ria@example.com",
                },
                {
                    "sponsor_name": "Ready Gold",
                    "active": "Check",
                    "package": "Gold",
                    "first_name": "Ben",
                    "to_email": "ben@example.com",
                },
            ]
        )
        self.addCleanup(fixture.cleanup)
        fixture.write_kontaktliste()
        fixture.write_gespraechsplan("Ready Premium_Gesprächsplan.pdf")
        fixture.write_gespraechsplan("Ready Gold_Gesprächsplan.pdf")
        fixture.write_vortragsliste("Ready Gold_Vortragsliste.pdf")

        result = fixture.assemble(
            kontaktliste_attachment_builder=lambda path, password: (_ for _ in ()).throw(RuntimeError("Excel COM kaputt"))
        )

        self.assertFalse(result.kontaktliste_protected)
        self.assertIn("Verschlüsselung fehlgeschlagen", result.kontaktliste_protection_details)
        self.assertEqual(0, result.ready_count)
        self.assertEqual(2, result.blocked_count)
        self.assertTrue(all(not sponsor.is_ready for sponsor in result.sponsors))

    def test_multiple_matching_gespraechsplaene_are_all_attached(self) -> None:
        fixture = _EventEndFixture(
            [
                {
                    "sponsor_name": "HP Deutschland",
                    "active": "Check",
                    "package": "Gold",
                    "first_name": "Hanna",
                    "to_email": "hanna@example.com",
                }
            ]
        )
        self.addCleanup(fixture.cleanup)
        fixture.write_kontaktliste()
        fixture.write_gespraechsplan("HP Deutschland Mergim Kambera_Gesprächsplan.pdf")
        fixture.write_gespraechsplan("HP Deutschland Michael Gieseke_Gesprächsplan.pdf")
        fixture.write_vortragsliste("HP Deutschland_Vortragsliste.pdf")

        result = fixture.assemble()

        sponsor = result.sponsors[0]
        self.assertTrue(sponsor.is_ready)
        self.assertEqual(
            (
                "Kontaktliste.xlsx",
                "HP Deutschland Mergim Kambera_Gesprächsplan.pdf",
                "HP Deutschland Michael Gieseke_Gesprächsplan.pdf",
                "HP Deutschland_Vortragsliste.pdf",
            ),
            sponsor.attachment_names,
        )

    def test_matching_handles_prefix_slash_and_truncated_names(self) -> None:
        fixture = _EventEndFixture(
            [
                {
                    "sponsor_name": "Hewlett Packard Enterprise",
                    "active": "Check",
                    "package": "Gold",
                    "first_name": "Hank",
                    "to_email": "hank@example.com",
                },
                {
                    "sponsor_name": "Palo Alto / IBM",
                    "active": "Check",
                    "package": "Gold",
                    "first_name": "Pia",
                    "to_email": "pia@example.com",
                },
                {
                    "sponsor_name": "Trend Micro / TrendAI",
                    "active": "Check",
                    "package": "Gold",
                    "first_name": "Tom",
                    "to_email": "tom@example.com",
                },
            ]
        )
        self.addCleanup(fixture.cleanup)
        fixture.write_kontaktliste()
        fixture.write_gespraechsplan("Hewlett Packard Enterpris_Gesprächsplan.pdf")
        fixture.write_vortragsliste("Hewlett Packard Enterpris_Vortragsliste.pdf")
        fixture.write_gespraechsplan("Palo Alto IBM_Gesprächsplan.pdf")
        fixture.write_vortragsliste("Palo Alto  IBM_Vortragsliste.pdf")
        fixture.write_gespraechsplan("Trend Micro_Gesprächsplan.pdf")
        fixture.write_vortragsliste("Trend Micro_Vortragsliste.pdf")

        result = fixture.assemble()
        sponsors = {sponsor.sponsor_name: sponsor for sponsor in result.sponsors}

        self.assertTrue(sponsors["Hewlett Packard Enterprise"].is_ready)
        self.assertIn("Hewlett Packard Enterpris_Gesprächsplan.pdf", sponsors["Hewlett Packard Enterprise"].attachment_names)
        self.assertTrue(sponsors["Palo Alto / IBM"].is_ready)
        self.assertIn("Palo Alto IBM_Gesprächsplan.pdf", sponsors["Palo Alto / IBM"].attachment_names)
        self.assertTrue(sponsors["Trend Micro / TrendAI"].is_ready)
        self.assertIn("Trend Micro_Vortragsliste.pdf", sponsors["Trend Micro / TrendAI"].attachment_names)

    def test_summary_dataframe_surfaces_readiness_and_attachments(self) -> None:
        fixture = _EventEndFixture(
            [
                {
                    "sponsor_name": "Ready Premium",
                    "active": "Check",
                    "package": "Premium",
                    "first_name": "Ria",
                    "to_email": "ria@example.com",
                },
                {
                    "sponsor_name": "Blocked Gold",
                    "active": "Check",
                    "package": "Gold",
                    "first_name": "Ben",
                    "to_email": "ben@example.com",
                },
            ]
        )
        self.addCleanup(fixture.cleanup)
        fixture.write_kontaktliste()
        fixture.write_gespraechsplan("Ready Premium_Gesprächsplan.pdf")

        result = fixture.assemble()
        summary_df = build_eventende_summary_dataframe(result.sponsors)

        self.assertEqual(["Bereit", "Blockiert"], summary_df["Status"].tolist())
        self.assertIn("Kontaktliste.xlsx", summary_df.loc[0, "Anhänge"])
        self.assertIn("Vortragsliste fehlt", summary_df.loc[1, "Hinweis"])


class BuildEventEndSerienmailsTests(unittest.TestCase):
    def test_builds_ready_serienmails_with_cc_and_attachments(self) -> None:
        fixture = _EventEndFixture(
            [
                {
                    "sponsor_name": "Acme Gold",
                    "active": "Check",
                    "package": "Gold",
                    "first_name": "Gina",
                    "to_email": "gina@example.com",
                    "cc_email": "copy@example.com",
                }
            ]
        )
        self.addCleanup(fixture.cleanup)
        fixture.write_kontaktliste()
        fixture.write_gespraechsplan("Acme Gold_Gesprächsplan.pdf")
        fixture.write_vortragsliste("Acme Gold_Vortragsliste.pdf")

        result = fixture.assemble()

        mails = build_eventende_serienmails(
            result.sponsors,
            subject_template="Unterlagen für {firma}",
            body_html_template="<p>Hallo {vorname}</p>",
            sender_email="sender@example.com",
        )

        self.assertEqual(1, len(mails))
        self.assertEqual("copy@example.com", mails[0].cc_email)
        self.assertEqual(3, len(mails[0].attachments))
        self.assertEqual("Unterlagen für Acme Gold", mails[0].subject)
